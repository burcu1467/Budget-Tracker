from fastapi import FastAPI, HTTPException, Header, Depends, status, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse, HTMLResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import json, os, sqlite3, shutil
import socket
import uvicorn
import time
from io import BytesIO
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from urllib.parse import unquote
import uuid
import sys
import re
import subprocess
from fastapi.security import OAuth2PasswordBearer

from passlib.context import CryptContext
from jose import JWTError, jwt
import multipart
from PIL import Image, ImageEnhance
import pytesseract
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# Pillow sürüm uyumluluğu
try:
    resample_filter = Image.Resampling.BICUBIC
except AttributeError:
    resample_filter = Image.BICUBIC

# Tesseract-OCR'nin kurulu olduğu yolu belirtin (Windows için).
if os.name == 'nt':
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
else:
    pytesseract.pytesseract.tesseract_cmd = 'tesseract'

app = FastAPI(title="Bütçe Takipçisi API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, "data")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")

# Data klasörü yoksa oluştur
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

# Uploads klasörü yoksa oluştur
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

DB_FILE = os.path.join(DATA_DIR, "butce.db")

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    # Kullanıcılar
    c.execute('''CREATE TABLE IF NOT EXISTS users 
                 (username TEXT PRIMARY KEY, password_hash TEXT, recovery_key_hash TEXT, email TEXT)''')
    
    # Kullanıcılar tablosu sütun kontrolü (Eski veritabanları için)
    c.execute("PRAGMA table_info(users)")
    columns = [col[1] for col in c.fetchall()]
    if "recovery_key_hash" not in columns:
        c.execute("ALTER TABLE users ADD COLUMN recovery_key_hash TEXT")
    if "email" not in columns:
        c.execute("ALTER TABLE users ADD COLUMN email TEXT")
    if "is_verified" not in columns:
        c.execute("ALTER TABLE users ADD COLUMN is_verified INTEGER DEFAULT 0")
    if "verification_token" not in columns:
        c.execute("ALTER TABLE users ADD COLUMN verification_token TEXT")

    # Harcamalar
    c.execute('''CREATE TABLE IF NOT EXISTS harcamalar 
                 (id TEXT PRIMARY KEY, username TEXT, aciklama TEXT, tutar REAL, kategori TEXT, tarih TEXT)''')
    # Harcamalar tablosuna fis_dosyasi sütunu ekle (Geriye uyumluluk)
    c.execute("PRAGMA table_info(harcamalar)")
    columns = [col[1] for col in c.fetchall()]
    if "fis_dosyasi" not in columns:
        c.execute("ALTER TABLE harcamalar ADD COLUMN fis_dosyasi TEXT")

    # Gelirler
    c.execute('''CREATE TABLE IF NOT EXISTS gelirler 
                 (id TEXT PRIMARY KEY, username TEXT, aciklama TEXT, tutar REAL, tarih TEXT)''')
    # Bütçe
    c.execute('''CREATE TABLE IF NOT EXISTS butce 
                 (username TEXT PRIMARY KEY, tutar REAL)''')
    # Tekrarlayan İşlemler
    c.execute('''CREATE TABLE IF NOT EXISTS tekrarlayan 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT, aciklama TEXT, tutar REAL, kategori TEXT, gun INTEGER, aktif BOOLEAN)''')
    # Kumbara Ayarları
    c.execute('''CREATE TABLE IF NOT EXISTS kumbara_ayarlar 
                 (username TEXT PRIMARY KEY, bakiye REAL DEFAULT 0, mod TEXT, gunluk_tutar REAL, haftalik_tutar REAL, son_tarih TEXT)''')
    # Kumbara İşlemleri
    c.execute('''CREATE TABLE IF NOT EXISTS kumbara_islemleri 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT, tarih TEXT, tutar REAL, tur TEXT, aciklama TEXT)''')
    # Kategori Bütçeleri
    c.execute('''CREATE TABLE IF NOT EXISTS kategori_butce 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT, kategori TEXT, limit_tutar REAL)''')
    # Özel Kategoriler
    c.execute('''CREATE TABLE IF NOT EXISTS custom_categories 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT, name TEXT)''')
    conn.commit()
    conn.close()

init_db()

def get_db():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

# --- Dil ve Mesajlar ---
MESSAGES = {
    "tr": {
        "invalid_username": "Geçersiz kullanıcı adı",
        "recovery_key_required": "Kurtarma anahtarı zorunludur",
        "username_taken": "Bu kullanıcı adı zaten alınmış",
        "registration_success": "Kayıt başarılı",
        "server_error": "Sunucu hatası: {}",
        "invalid_username_format": "Geçersiz kullanıcı adı formatı",
        "invalid_credentials": "Kullanıcı adı veya şifre hatalı",
        "session_expired": "Oturum süresi doldu veya geçersiz",
        "recovery_key_not_found": "Bu kullanıcı için kurtarma anahtarı bulunamadı",
        "recovery_key_invalid": "Kurtarma anahtarı hatalı",
        "password_reset_success": "Şifre başarıyla sıfırlandı",
        "user_not_found": "Kullanıcı bulunamadı",
        "current_password_invalid": "Mevcut şifre hatalı",
        "recovery_key_updated": "Kurtarma anahtarı güncellendi",
        "password_change_success": "Şifre başarıyla değiştirildi",
        "expense_not_found": "Harcama bulunamadı",
        "income_not_found": "Gelir bulunamadı",
        "email_required": "Önce profilinizden bir e-posta adresi kaydedin.",
        "already_verified": "Zaten doğrulanmış.",
        "verification_sent": "Doğrulama bağlantısı sunucu konsoluna yazdırıldı (Simülasyon).",
        "insufficient_balance": "Yetersiz bakiye"
    },
    "en": {
        "invalid_username": "Invalid username",
        "recovery_key_required": "Recovery key is required",
        "username_taken": "Username already taken",
        "registration_success": "Registration successful",
        "server_error": "Server error: {}",
        "invalid_username_format": "Invalid username format",
        "invalid_credentials": "Invalid username or password",
        "session_expired": "Session expired or invalid",
        "recovery_key_not_found": "Recovery key not found for this user",
        "recovery_key_invalid": "Invalid recovery key",
        "password_reset_success": "Password reset successfully",
        "user_not_found": "User not found",
        "current_password_invalid": "Invalid current password",
        "recovery_key_updated": "Recovery key updated",
        "password_change_success": "Password changed successfully",
        "expense_not_found": "Expense not found",
        "income_not_found": "Income not found",
        "email_required": "Please save an email address in your profile first.",
        "already_verified": "Already verified.",
        "verification_sent": "Verification link printed to server console (Simulation).",
        "insufficient_balance": "Insufficient balance"
    }
}

CATEGORIES = {
    "tr": ["Yemek", "Ulaşım", "Eğlence", "Alışveriş", "Sağlık", "Faturalar", "Diğer"],
    "en": ["Food", "Transportation", "Entertainment", "Shopping", "Health", "Bills", "Other"]
}

def get_language(accept_language: str = Header(default="tr", alias="Accept-Language")):
    if not accept_language:
        return "tr"
    # "en-US,en;q=0.9" gibi başlıkları ayrıştır
    lang = accept_language.split(",")[0].split("-")[0].lower()
    return lang if lang in MESSAGES else "tr"

def get_msg(key: str, lang: str = "tr", *args):
    msg = MESSAGES.get(lang, MESSAGES["tr"]).get(key, key)
    if args:
        return msg.format(*args)
    return msg

# --- Güvenlik ve Parola Yönetimi ---
# bcrypt: yeni ve güvenli hash algoritması
# plain_hex_sha256: eski, salt'suz sha256 hash'leri ile geriye dönük uyumluluk için
# 'deprecated' olarak işaretlenmesi, eski formatta bir parola doğrulandığında
# otomatik olarak bcrypt'e yükseltilmesini sağlar.
pwd_context = CryptContext(
    schemes=["pbkdf2_sha256", "bcrypt"],
    deprecated="auto"
)

def get_hash(text: str) -> str:
    return pwd_context.hash(text)

# --- JWT Ayarları ---
SECRET_KEY = os.getenv("SECRET_KEY", "cok-gizli-ve-guvenli-bir-anahtar-buraya-yazilmali") # Prodüksiyonda env'den alınmalı
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7 # 1 Hafta

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="api/login")

def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def get_current_user(token: str = Depends(oauth2_scheme), lang: str = Depends(get_language)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail=get_msg("session_expired", lang),
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    return {"username": username, "lang": lang}

class UserAuth(BaseModel):
    username: str
    password: str
    recovery_key: str = None

class HarcamaIn(BaseModel):
    aciklama: str
    tutar: float
    kategori: str
    tarih: str = None

class GelirIn(BaseModel):
    aciklama: str
    tutar: float
    tarih: str = None

class ButceIn(BaseModel):
    tutar: float

class KategoriButceIn(BaseModel):
    kategori: str
    limit_tutar: float

class CategoryIn(BaseModel):
    name: str

class TekrarlayanIn(BaseModel):
    aciklama: str
    tutar: float
    kategori: str
    gun: int
    aktif: bool = True

class KumbaraAyarIn(BaseModel):
    mod: str = None # "gunluk", "haftalik"
    gunluk_tutar: float = 0
    haftalik_tutar: float = 0

class KumbaraIslemIn(BaseModel):
    tutar: float
    tur: str # "ekle", "cek"

class RecoveryUpdateIn(BaseModel):
    username: str
    password: str
    new_recovery_key: str

class PasswordChangeIn(BaseModel):
    username: str
    old_password: str
    new_password: str

class ProfileIn(BaseModel):
    email: str

def sync_user_data(username: str, conn: sqlite3.Connection):
    """Tekrarlayan harcamaları ve kumbara otomatik eklemelerini kontrol eder"""
    username = unquote(username).lower()
    c = conn.cursor()

    # 1. Tekrarlayan Harcamalar
    c.execute("SELECT * FROM tekrarlayan WHERE username = ? AND aktif = 1", (username,))
    tekrarlayan = c.fetchall()
    
    bugun = datetime.now()
    gun = bugun.day
    bugune_git = bugun.strftime("%d.%m.%Y")
    
    for t in tekrarlayan:
        if t['gun'] == gun:
            # Bugün için zaten eklenmiş mi?
            c.execute("SELECT 1 FROM harcamalar WHERE username=? AND tarih=? AND aciklama=?", 
                      (username, bugune_git, t['aciklama']))
            if not c.fetchone():
                new_id = str(uuid.uuid4())
                c.execute("INSERT INTO harcamalar (id, username, aciklama, tutar, kategori, tarih) VALUES (?, ?, ?, ?, ?, ?)",
                          (new_id, username, t['aciklama'], t['tutar'], t['kategori'], bugune_git))

    # 2. Kumbara Otomatik Ekleme
    c.execute("SELECT * FROM kumbara_ayarlar WHERE username = ?", (username,))
    ayar = c.fetchone()
    
    if ayar and ayar['mod']:
        son_tarih_str = ayar['son_tarih']
        bugun = datetime.now()
        son_tarih = datetime.strptime(son_tarih_str, "%d.%m.%Y") if son_tarih_str else None
        tutar_ekle = 0
        
        if ayar['mod'] == "gunluk":
            diff = (bugun - son_tarih).days if son_tarih else 1
            if diff > 0: tutar_ekle = ayar['gunluk_tutar'] * diff
        elif ayar['mod'] == "haftalik":
            diff = (bugun - son_tarih).days // 7 if son_tarih else 1
            if diff > 0: tutar_ekle = ayar['haftalik_tutar'] * diff
            
        if tutar_ekle > 0:
            new_bakiye = ayar['bakiye'] + tutar_ekle
            bugun_str = bugun.strftime("%d.%m.%Y")
            
            c.execute("UPDATE kumbara_ayarlar SET bakiye = ?, son_tarih = ? WHERE username = ?", 
                      (new_bakiye, bugun_str, username))
            c.execute("INSERT INTO kumbara_islemleri (username, tarih, tutar, tur, aciklama) VALUES (?, ?, ?, ?, ?)",
                      (username, bugun_str, tutar_ekle, "ekleme", "Otomatik"))
    
    conn.commit()

@app.post("/api/register")
def register(user: UserAuth, lang: str = Depends(get_language), conn: sqlite3.Connection = Depends(get_db)):
    try:
        safe_user = "".join(c for c in user.username if c.isalnum())
        if not safe_user:
            raise HTTPException(status_code=400, detail=get_msg("invalid_username", lang))
        
        if not user.recovery_key:
            raise HTTPException(status_code=400, detail=get_msg("recovery_key_required", lang))
        
        username_key = safe_user.lower()
        c = conn.cursor()
        
        c.execute("SELECT 1 FROM users WHERE username = ?", (username_key,))
        if c.fetchone():
            raise HTTPException(status_code=400, detail=get_msg("username_taken", lang))

        pw_hash = get_hash(user.password)
        rc_hash = get_hash(user.recovery_key)
        
        c.execute("INSERT INTO users (username, password_hash, recovery_key_hash) VALUES (?, ?, ?)", (username_key, pw_hash, rc_hash))
        conn.commit()
        
        return {"ok": True, "message": get_msg("registration_success", lang)}
    except Exception as e:
        print(f"❌ KAYIT HATASI: {e}")
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=get_msg("server_error", lang, str(e)))

@app.post("/api/login")
def login(user: UserAuth, lang: str = Depends(get_language), conn: sqlite3.Connection = Depends(get_db)):
    safe_user = "".join(c for c in user.username if c.isalnum())
    if not safe_user:
        print("❌ GİRİŞ BAŞARISIZ: Geçersiz kullanıcı adı formatı.")
        raise HTTPException(status_code=400, detail=get_msg("invalid_username_format", lang))
    
    username_key = safe_user.lower()
    c = conn.cursor()
    c.execute("SELECT password_hash FROM users WHERE username = ?", (username_key,))
    row = c.fetchone()
    
    if not row:
        print(f"❌ GİRİŞ BAŞARISIZ: '{username_key}' bulunamadı.")
        raise HTTPException(status_code=400, detail=get_msg("invalid_credentials", lang))

    verified, new_hash = pwd_context.verify_and_update(user.password, row['password_hash'])
    if not verified:
        print(f"❌ GİRİŞ BAŞARISIZ: '{username_key}' için şifre yanlış.")
        raise HTTPException(status_code=400, detail=get_msg("invalid_credentials", lang))
    
    if new_hash:
        c.execute("UPDATE users SET password_hash = ? WHERE username = ?", (new_hash, username_key))
        conn.commit()
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(data={"sub": username_key}, expires_delta=access_token_expires)
    
    return {"access_token": access_token, "token_type": "bearer", "username": username_key}

@app.post("/api/reset-password")
def reset_password(user: UserAuth, lang: str = Depends(get_language), conn: sqlite3.Connection = Depends(get_db)):
    safe_user = "".join(c for c in user.username if c.isalnum())
    if not safe_user:
        raise HTTPException(status_code=400, detail=get_msg("invalid_username", lang))
        
    username_key = safe_user.lower()
    
    c = conn.cursor()
    c.execute("SELECT recovery_key_hash FROM users WHERE username = ?", (username_key,))
    row = c.fetchone()
    
    if not row:
        raise HTTPException(status_code=400, detail=get_msg("recovery_key_not_found", lang))
        
    verified, new_hash = pwd_context.verify_and_update(user.recovery_key, row['recovery_key_hash'])
    if not verified:
        raise HTTPException(status_code=400, detail=get_msg("recovery_key_invalid", lang))
    
    if new_hash:
        # Kurtarma anahtarının hash'ini de güncelle
        c.execute("UPDATE users SET recovery_key_hash = ? WHERE username = ?", (new_hash, username_key))

    # Yeni şifreyi hash'le
    new_pw_hash = get_hash(user.password)
    c.execute("UPDATE users SET password_hash = ? WHERE username = ?", (new_pw_hash, username_key))
    conn.commit()
    return {"ok": True, "message": get_msg("password_reset_success", lang)}

@app.post("/api/update-recovery-key")
def update_recovery_key(data: RecoveryUpdateIn, lang: str = Depends(get_language), conn: sqlite3.Connection = Depends(get_db)):
    safe_user = "".join(c for c in data.username if c.isalnum())
    if not safe_user:
        raise HTTPException(status_code=400, detail=get_msg("invalid_username", lang))
        
    username_key = safe_user.lower()
    
    c = conn.cursor()
    c.execute("SELECT password_hash FROM users WHERE username = ?", (username_key,))
    row = c.fetchone()
    
    if not row:
        raise HTTPException(status_code=400, detail=get_msg("user_not_found", lang))

    verified, new_hash = pwd_context.verify_and_update(data.password, row['password_hash'])
    if not verified:
        raise HTTPException(status_code=400, detail=get_msg("current_password_invalid", lang))
    if new_hash:
        c.execute("UPDATE users SET password_hash = ? WHERE username = ?", (new_hash, username_key))
        
    rc_hash = get_hash(data.new_recovery_key)
    c.execute("UPDATE users SET recovery_key_hash = ? WHERE username = ?", (rc_hash, username_key))
    conn.commit()
    
    return {"ok": True, "message": get_msg("recovery_key_updated", lang)}

@app.post("/api/change-password")
def change_password(data: PasswordChangeIn, lang: str = Depends(get_language), conn: sqlite3.Connection = Depends(get_db)):
    safe_user = "".join(c for c in data.username if c.isalnum())
    if not safe_user:
        raise HTTPException(status_code=400, detail=get_msg("invalid_username", lang))
        
    username_key = safe_user.lower()
    
    c = conn.cursor()
    c.execute("SELECT password_hash FROM users WHERE username = ?", (username_key,))
    row = c.fetchone()
    
    if not row:
        raise HTTPException(status_code=400, detail=get_msg("user_not_found", lang))

    verified, new_hash = pwd_context.verify_and_update(data.old_password, row['password_hash'])
    if not verified:
        raise HTTPException(status_code=400, detail=get_msg("current_password_invalid", lang))
        
    # Eğer eski hash formatı kullanılıyorsa new_hash dolu gelir, ama biz zaten yeni şifre set edeceğimiz için bunu kullanmıyoruz.
    new_pw_hash = get_hash(data.new_password)
    c.execute("UPDATE users SET password_hash = ? WHERE username = ?", (new_pw_hash, username_key))
    conn.commit()
    
    return {"ok": True, "message": get_msg("password_change_success", lang)}
@app.get("/api/profile")
def get_profile(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("SELECT email, is_verified FROM users WHERE username = ?", (username,))
    row = c.fetchone()
    return {
        "email": row['email'] if row else "",
        "is_verified": row['is_verified'] if row else False
    }

@app.post("/api/profile")
def update_profile(p: ProfileIn, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    # E-posta değişirse doğrulama durumunu sıfırla
    c.execute("SELECT email FROM users WHERE username = ?", (username,))
    row = c.fetchone()
    if row and row['email'] != p.email:
        c.execute("UPDATE users SET email = ?, is_verified = 0, verification_token = NULL WHERE username = ?", (p.email, username))
    else:
        c.execute("UPDATE users SET email = ? WHERE username = ?", (p.email, username))
    conn.commit()
    return {"ok": True}

@app.get("/api/categories")
def get_categories(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    lang = current_user['lang']
    
    defaults = CATEGORIES.get(lang, CATEGORIES["tr"])
    
    c = conn.cursor()
    c.execute("SELECT name FROM custom_categories WHERE username = ?", (username,))
    customs = [row['name'] for row in c.fetchall()]
    
    combined = list(defaults)
    for cat in customs:
        if cat not in combined:
            combined.append(cat)
            
    return combined

@app.post("/api/categories")
def add_category(cat: CategoryIn, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("SELECT 1 FROM custom_categories WHERE username = ? AND name = ?", (username, cat.name))
    if c.fetchone():
        raise HTTPException(status_code=400, detail="Kategori zaten mevcut")
    c.execute("INSERT INTO custom_categories (username, name) VALUES (?, ?)", (username, cat.name))
    conn.commit()
    return {"ok": True}

@app.put("/api/categories/{old_name}")
def update_category(old_name: str, cat: CategoryIn, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("UPDATE custom_categories SET name = ? WHERE username = ? AND name = ?", (cat.name, username, old_name))
    if c.rowcount == 0:
        raise HTTPException(status_code=404, detail="Düzenlenecek özel kategori bulunamadı (Varsayılanlar düzenlenemez)")
    conn.commit()
    return {"ok": True}

@app.delete("/api/categories/{name}")
def delete_category(name: str, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("DELETE FROM custom_categories WHERE username = ? AND name = ?", (username, name))
    if c.rowcount == 0:
        raise HTTPException(status_code=404, detail="Silinecek özel kategori bulunamadı (Varsayılanlar silinemez)")
    conn.commit()
    return {"ok": True}

@app.get("/api/harcamalar")
def list_harcamalar(kategori: str = None, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    if kategori:
        c.execute("SELECT * FROM harcamalar WHERE username = ? AND kategori = ? ORDER BY rowid", (username, kategori))
    else:
        c.execute("SELECT * FROM harcamalar WHERE username = ? ORDER BY rowid", (username,))
    rows = [dict(row) for row in c.fetchall()]
    return rows

@app.post("/api/harcamalar")
def add_harcama(
    aciklama: str = Form(...),
    tutar: float = Form(...),
    kategori: str = Form(...),
    tarih: str = Form(None),
    fis_dosyasi: UploadFile = File(None),
    current_user: dict = Depends(get_current_user), 
    conn: sqlite3.Connection = Depends(get_db)
):
    username = current_user['username']
    
    item_tarih = tarih if tarih and tarih.strip() else datetime.now().strftime("%d.%m.%Y")
    item_id = str(uuid.uuid4())
    file_path_to_db = None

    if fis_dosyasi:
        # Güvenli bir dosya adı oluştur
        file_extension = os.path.splitext(fis_dosyasi.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_save_path = os.path.join(UPLOAD_DIR, unique_filename)
        
        # Dosyayı diske kaydet
        try:
            with open(file_save_path, "wb") as buffer:
                shutil.copyfileobj(fis_dosyasi.file, buffer)
            file_path_to_db = unique_filename
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Dosya kaydedilemedi: {e}")
        finally:
            fis_dosyasi.file.close()

    c = conn.cursor()
    c.execute("INSERT INTO harcamalar (id, username, aciklama, tutar, kategori, tarih, fis_dosyasi) VALUES (?, ?, ?, ?, ?, ?, ?)",
              (item_id, username, aciklama, tutar, kategori, item_tarih, file_path_to_db))
    conn.commit()
    
    # Güncel listeyi dön
    c.execute("SELECT * FROM harcamalar WHERE username = ? ORDER BY rowid", (username,))
    rows = [dict(row) for row in c.fetchall()]
    return rows

@app.delete("/api/harcamalar/{identifier}")
def delete_harcama(identifier: str, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    lang = current_user['lang']    
    c = conn.cursor()
    
    # 1. ID'ye göre silmeyi dene
    c.execute("SELECT id, fis_dosyasi FROM harcamalar WHERE id = ? AND username = ?", (identifier, username))
    row = c.fetchone()
    
    target_id = None
    file_to_delete = None

    if row:
        target_id = row['id']
        file_to_delete = row['fis_dosyasi']
    elif identifier.isdigit():
        # 2. ID değilse ve sayıysa (index), eski usul silmeyi dene
        c.execute("SELECT id, fis_dosyasi FROM harcamalar WHERE username = ? ORDER BY rowid", (username,))
        rows = c.fetchall()
        index = int(identifier)
        if 0 <= index < len(rows):
            target_id = rows[index]['id']
            file_to_delete = rows[index]['fis_dosyasi']

    if target_id:
        c.execute("DELETE FROM harcamalar WHERE id = ?", (target_id,))
        conn.commit()
        if file_to_delete:
            file_path = os.path.join(UPLOAD_DIR, file_to_delete)
            # Dosya varsa sil, hata olursa yoksay
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except OSError:
                    pass
        return {"ok": True}
            
    raise HTTPException(status_code=404, detail=get_msg("expense_not_found", lang))

@app.get("/api/gelirler")
def list_gelirler(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("SELECT * FROM gelirler WHERE username = ? ORDER BY rowid", (username,))
    rows = [dict(row) for row in c.fetchall()]
    return rows

@app.post("/api/gelirler")
def add_gelir(g: GelirIn, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    item = g.dict()
    if not item.get("tarih") or not item["tarih"].strip():
        item["tarih"] = datetime.now().strftime("%d.%m.%Y")
    item["id"] = str(uuid.uuid4())
    
    c = conn.cursor()
    c.execute("INSERT INTO gelirler (id, username, aciklama, tutar, tarih) VALUES (?, ?, ?, ?, ?)",
              (item["id"], username, item["aciklama"], item["tutar"], item["tarih"]))
    conn.commit()
    
    c.execute("SELECT * FROM gelirler WHERE username = ?", (username,))
    rows = [dict(row) for row in c.fetchall()]
    return rows

@app.delete("/api/gelirler/{identifier}")
def delete_gelir(identifier: str, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    lang = current_user['lang']
    c = conn.cursor()
    
    c.execute("DELETE FROM gelirler WHERE id = ? AND username = ?", (identifier, username))
    if c.rowcount > 0:
        conn.commit()
        return {"ok": True}
            
    if identifier.isdigit():
        # DİKKAT: Sıralama (ORDER BY) eklenerek silme işleminin tutarlı olması sağlandı.
        c.execute("SELECT id FROM gelirler WHERE username = ? ORDER BY rowid", (username,))
        rows = c.fetchall()
        index = int(identifier)
        if 0 <= index < len(rows):
            target_id = rows[index]['id']
            c.execute("DELETE FROM gelirler WHERE id = ?", (target_id,))
            conn.commit()
            return {"ok": True}

    raise HTTPException(status_code=404, detail=get_msg("income_not_found", lang))

@app.get("/api/ozet")
def ozet(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    sync_user_data(username, conn) # Verileri güncelle (artık bağlantıyı dışarıdan alıyor)
    c = conn.cursor()
    
    c.execute("SELECT SUM(tutar) as total FROM harcamalar WHERE username = ?", (username,))
    res_gider = c.fetchone()
    toplam_gider = res_gider['total'] if res_gider and res_gider['total'] else 0
    
    c.execute("SELECT SUM(tutar) as total FROM gelirler WHERE username = ?", (username,))
    res_gelir = c.fetchone()
    toplam_gelir = res_gelir['total'] if res_gelir and res_gelir['total'] else 0
    
    c.execute("SELECT tutar FROM butce WHERE username = ?", (username,))
    res_butce = c.fetchone()
    butce_val = res_butce['tutar'] if res_butce else 0
    
    return {
        "toplam_gider": toplam_gider, 
        "toplam_gelir": toplam_gelir, 
        "net": toplam_gelir - toplam_gider,
        "butce": butce_val,
        "kalan_butce": butce_val - toplam_gider
    }

@app.get("/uploads/{filename}")
async def get_upload(filename: str, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    # Güvenlik: Kullanıcının sadece kendi harcamalarına ait dosyaları görebilmesini sağla
    username = current_user['username']
    c = conn.cursor()
    c.execute("SELECT 1 FROM harcamalar WHERE username = ? AND fis_dosyasi = ?", (username, filename))
    if not c.fetchone():
        raise HTTPException(status_code=404, detail="Dosya bulunamadı veya erişim yetkiniz yok.")

    file_path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="Dosya bulunamadı.")
    
    # Path traversal saldırılarını önlemek için ek kontrol
    if not os.path.abspath(file_path).startswith(os.path.abspath(UPLOAD_DIR)):
        raise HTTPException(status_code=403, detail="Geçersiz dosya yolu.")

    return FileResponse(file_path)
@app.get("/api/verify-email")
def verify_email(token: str, conn: sqlite3.Connection = Depends(get_db)):
    c = conn.cursor()
    c.execute("SELECT username FROM users WHERE verification_token = ?", (token,))
    row = c.fetchone()
    
    if not row:
        return HTMLResponse(content="<h1 style='color:red; font-family:sans-serif; text-align:center; margin-top:50px;'>Geçersiz veya süresi dolmuş doğrulama linki.</h1>", status_code=400)
    
    username = row['username']
    c.execute("UPDATE users SET is_verified = 1, verification_token = NULL WHERE username = ?", (username,))
    conn.commit()
    return HTMLResponse(content="<h1 style='color:green; font-family:sans-serif; text-align:center; margin-top:50px;'>✅ E-posta başarıyla doğrulandı!</h1><p style='text-align:center; font-family:sans-serif;'>Uygulamaya geri dönebilirsiniz.</p>")

@app.post("/api/resend-verification")
def resend_verification(request: Request, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    lang = current_user['lang']
    c = conn.cursor()
    c.execute("SELECT email, is_verified FROM users WHERE username = ?", (username,))
    row = c.fetchone()
    
    if not row or not row['email']:
        raise HTTPException(status_code=400, detail=get_msg("email_required", lang))
        
    if row['is_verified']:
        return {"ok": False, "message": get_msg("already_verified", lang)}
    
    new_token = str(uuid.uuid4())
    c.execute("UPDATE users SET verification_token = ? WHERE username = ?", (new_token, username))
    conn.commit()
    
    # Gerçek e-posta gönderimi yerine konsola yazdırıyoruz
    print(f"\n📧 [SİMÜLASYON] Doğrulama Linki: {request.base_url}api/verify-email?token={new_token}\n")
    return {"ok": True, "message": get_msg("verification_sent", lang)}

# --- YENİ EKLENEN ENDPOINTLER ---

@app.get("/api/butce")
def get_butce(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("SELECT tutar FROM butce WHERE username = ?", (username,))
    row = c.fetchone()
    return {"butce": row['tutar'] if row else 0}

@app.post("/api/butce")
def set_butce(b: ButceIn, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO butce (username, tutar) VALUES (?, ?)", (username, b.tutar))
    conn.commit()
    return {"ok": True, "butce": b.tutar}

@app.get("/api/butce/kategori")
def get_kategori_butce(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("SELECT kategori, limit_tutar FROM kategori_butce WHERE username = ?", (username,))
    rows = [dict(row) for row in c.fetchall()]
    return rows
@app.post("/api/butce/kategori")
def set_kategori_butce(kb: KategoriButceIn, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("DELETE FROM kategori_butce WHERE username = ? AND kategori = ?", (username, kb.kategori))
    if kb.limit_tutar > 0:
        c.execute("INSERT INTO kategori_butce (username, kategori, limit_tutar) VALUES (?, ?, ?)", (username, kb.kategori, kb.limit_tutar))
    conn.commit()
    return {"ok": True}

@app.get("/api/tekrarlayan")
def list_tekrarlayan(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("SELECT * FROM tekrarlayan WHERE username = ? ORDER BY id", (username,))
    rows = [dict(row) for row in c.fetchall()]
    return rows

@app.post("/api/tekrarlayan")
def add_tekrarlayan(t: TekrarlayanIn, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("INSERT INTO tekrarlayan (username, aciklama, tutar, kategori, gun, aktif) VALUES (?, ?, ?, ?, ?, ?)",
              (username, t.aciklama, t.tutar, t.kategori, t.gun, 1 if t.aktif else 0))
    conn.commit()
    return {"ok": True}

@app.delete("/api/tekrarlayan/{index}")
def delete_tekrarlayan(index: int, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    # Frontend index gönderdiği için önce tüm listeyi çekip ID'yi buluyoruz
    c.execute("SELECT id FROM tekrarlayan WHERE username = ? ORDER BY id", (username,))
    rows = c.fetchall()
    
    if 0 <= index < len(rows):
        target_id = rows[index]['id']
        c.execute("DELETE FROM tekrarlayan WHERE id = ?", (target_id,))
        conn.commit()
        return {"ok": True}
        
    raise HTTPException(status_code=404)

@app.get("/api/kumbara")
def get_kumbara(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    
    c.execute("SELECT * FROM kumbara_ayarlar WHERE username = ?", (username,))
    ayar_row = c.fetchone()
    
    c.execute("SELECT * FROM kumbara_islemleri WHERE username = ?", (username,))
    islemler = [dict(row) for row in c.fetchall()]
    
    
    ayar = dict(ayar_row) if ayar_row else {"bakiye": 0, "mod": None, "gunluk_tutar": 0, "haftalik_tutar": 0, "son_tarih": None}
    
    return {
        "bakiye": ayar.get("bakiye", 0),
        "islemler": islemler,
        "ayar": ayar
    }

@app.post("/api/kumbara/ayarlar")
def set_kumbara_ayar(a: KumbaraAyarIn, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    son_tarih = datetime.now().strftime("%d.%m.%Y")
    c = conn.cursor()
    
    # Mevcut bakiye varsa koru, yoksa 0
    c.execute("SELECT bakiye FROM kumbara_ayarlar WHERE username = ?", (username,))
    row = c.fetchone()
    bakiye = row['bakiye'] if row else 0
    
    c.execute("INSERT OR REPLACE INTO kumbara_ayarlar (username, bakiye, mod, gunluk_tutar, haftalik_tutar, son_tarih) VALUES (?, ?, ?, ?, ?, ?)",
              (username, bakiye, a.mod, a.gunluk_tutar, a.haftalik_tutar, son_tarih))
    conn.commit()
    return {"ok": True}

@app.post("/api/kumbara/islem")
def kumbara_islem(i: KumbaraIslemIn, current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    lang = current_user['lang']
    c = conn.cursor()

    c.execute("SELECT * FROM kumbara_ayarlar WHERE username = ?", (username,))
    ayar_row = c.fetchone()
    bakiye = ayar_row['bakiye'] if ayar_row else 0

    if i.tur == "ekle":
        bakiye += i.tutar
    elif i.tur == "cek":
        if bakiye < i.tutar:
            raise HTTPException(status_code=400, detail=get_msg("insufficient_balance", lang))
        bakiye -= i.tutar
    
    tarih = datetime.now().strftime("%d.%m.%Y")
    
    # Ayarları güncelle (bakiye için) - Eğer ayar yoksa oluştur, varsa güncelle
    if ayar_row:
        c.execute("UPDATE kumbara_ayarlar SET bakiye = ? WHERE username = ?", (bakiye, username))
    else:
        c.execute("INSERT INTO kumbara_ayarlar (username, bakiye) VALUES (?, ?)", (username, bakiye))

    # İşlem ekle
    c.execute("INSERT INTO kumbara_islemleri (username, tarih, tutar, tur, aciklama) VALUES (?, ?, ?, ?, ?)",
              (username, tarih, i.tutar, i.tur, "Manuel işlem"))
    
    conn.commit()
    return {"ok": True, "bakiye": bakiye}

def parse_ocr_amount(val_str):
    """OCR sayısal değerlerini akıllıca ayrıştırır (TR/EN formatı)"""
    try:
        val_str = val_str.replace(' ', '').strip()
        val_str = re.sub(r'[^\d.,]', '', val_str)
        if not val_str: return None
        
        if '.' in val_str and ',' in val_str:
            if val_str.rfind(',') > val_str.rfind('.'):
                return float(val_str.replace('.', '').replace(',', '.')) # 1.234,56
            else:
                return float(val_str.replace(',', '')) # 1,234.56
        elif ',' in val_str:
            return float(val_str.replace(',', '.')) # 123,45
        elif '.' in val_str:
            parts = val_str.split('.')
            if len(parts) > 2: return float(val_str.replace('.', '')) # 1.234.567
            if len(parts[-1]) == 3: return float(val_str.replace('.', '')) # 1.234 -> 1234
            return float(val_str) # 123.45
        return float(val_str)
    except:
        return None

@app.post("/api/ocr")
def ocr_scan(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """
    Yüklenen fiş/fatura görselinden metin okur ve veri ayıklar.
    Gereksinim: Sunucuda Tesseract-OCR kurulu olmalıdır.
    NOT: CPU yoğun işlem olduğu için 'async def' yerine 'def' kullanıldı (Thread Pool).
    """
    try:
        # Senkron okuma (Thread içinde güvenli)
        contents = file.file.read()
        image = Image.open(BytesIO(contents))
        
        # OCR başarısını artırmak için görseli siyah-beyaz (grayscale) yap
        image = image.convert('L')
        
        # Görüntü iyileştirme (Boyutlandırma ve Kontrast)
        # Fişlerdeki küçük yazıları (özellikle 1 ve 7 ayrımı) okumak için büyütme ve netleştirme
        image = image.resize((image.width * 2, image.height * 2), resample_filter)
        
        # Keskinleştirme (Virgül ve noktaları korumak için)
        enhancer_sharp = ImageEnhance.Sharpness(image)
        image = enhancer_sharp.enhance(2.0)
        # Kontrast (Daha yumuşak artış)
        enhancer_cont = ImageEnhance.Contrast(image)
        image = enhancer_cont.enhance(1.5)

        # Türkçe ve İngilizce tarama yapmayı dene
        try:
            text = pytesseract.image_to_string(image, lang='tur+eng')
        except pytesseract.TesseractError:
            # Dil paketi yoksa varsayılan ile dene
            text = pytesseract.image_to_string(image)
        
        # Basit veri ayıklama (Regex)
        data = {
            "text": text,
            "tarih": None,
            "toplam": None,
            "isyeri": None
        }
        
        # 1. Tarih Bulma (Esnek Regex: GG.AA.YYYY, GG.AA.YY vb.)
        tarih_match = re.search(r'(\d{1,2})\s*[./-]\s*(\d{1,2})\s*[./-]\s*(\d{2,4})', text)
        if tarih_match:
            day, month, year = tarih_match.groups()
            if len(year) == 2: year = "20" + year
            data["tarih"] = f"{day.zfill(2)}.{month.zfill(2)}.{year}"
        else:
            # YYYY-MM-DD formatı
            tarih_match_2 = re.search(r'(\d{4})\s*[./-]\s*(\d{1,2})\s*[./-]\s*(\d{1,2})', text)
            if tarih_match_2:
                year, month, day = tarih_match_2.groups()
                data["tarih"] = f"{day.zfill(2)}.{month.zfill(2)}.{year}"
            
        # 2. Toplam Tutar Bulma
        # Tüm olası tutarları bul
        all_amounts = []
        # Regex: En az bir ayraç içeren veya içermeyen sayılar. Sonunda 2 hane olma zorunluluğunu esnettik.
        matches = re.findall(r'(?:\b|[^0-9])(\d+(?:\s*[.,]\s*\d+)*)(?:\b|[^0-9])', text)
        for m in matches:
            val = parse_ocr_amount(m)
            if val is not None and val < 1000000: # Mantıksız büyük sayıları ele
                # Yıl olma ihtimali olan sayıları ele (Örn: 2023, 2024)
                if 2000 <= val <= 2100 and val % 1 == 0:
                    continue
                # Eğer bulunan sayı, fişin yılına eşitse (Örn: 2025), bunu tutar sanma
                if data["tarih"] and str(int(val)) in data["tarih"]:
                     # Basit bir kontrol: Tutar tam sayı ise ve tarihin içinde geçiyorsa (Yıl olma ihtimali)
                     if val % 1 == 0 and len(str(int(val))) == 4: continue
                all_amounts.append(val)

        lines = text.split('\n')
        found_amount = None
        
        for i, line in enumerate(lines):
            line_upper = line.upper()
            if any(k in line_upper for k in ["TOPLAM", "TUTAR", "TOP", "GENEL", "ODENECEK", "ODENEN", "KREDİ"]):
                # Satırdaki sayıları kontrol et
                line_matches = re.findall(r'(\d+(?:\s*[.,]\s*\d+)*)', line)
                valid_line_vals = []
                for m in line_matches:
                    val = parse_ocr_amount(m)
                    if val is not None and val < 1000000:
                        valid_line_vals.append(val)
                
                if valid_line_vals:
                    found_amount = max(valid_line_vals)
                    break
                
                # Alt satıra bak
                if i + 1 < len(lines):
                    next_matches = re.findall(r'(\d+(?:\s*[.,]\s*\d+)*)', lines[i+1])
                    valid_next_vals = []
                    for m in next_matches:
                        val = parse_ocr_amount(m)
                        if val is not None and val < 1000000:
                            valid_next_vals.append(val)
                    
                    if valid_next_vals:
                        found_amount = max(valid_next_vals)
                        break
        
        if found_amount:
            data["toplam"] = found_amount
        elif all_amounts:
            # Anahtar kelime yoksa en büyük sayıyı al
            data["toplam"] = max(all_amounts)
        
        return {"ok": True, "data": data}
        
    except Exception as e:
        print(f"OCR Hatası: {e}")
        raise HTTPException(status_code=500, detail=f"OCR işlemi başarısız. Tesseract yüklü mü? Hata: {str(e)}")

@app.post("/api/harcamalar/otomatik-fis")
def add_harcama_ocr(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    conn: sqlite3.Connection = Depends(get_db)
):
    """
    Yüklenen fişi OCR ile tarar, verileri ayıklar ve otomatik olarak harcamalara ekler.
    """
    username = current_user['username']
    
    # 1. Dosyayı kaydet
    file_extension = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    file_save_path = os.path.join(UPLOAD_DIR, unique_filename)
    
    try:
        with open(file_save_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Dosya kaydedilemedi: {e}")
    finally:
        file.file.close()

    # 2. OCR ve Veri Çıkarımı
    aciklama = "Fiş Taraması"
    tutar = 0.0
    tarih = datetime.now().strftime("%d.%m.%Y")
    
    try:
        image = Image.open(file_save_path)
        # OCR başarısını artırmak için görseli siyah-beyaz (grayscale) yap
        image = image.convert('L')
        
        # Görüntü iyileştirme (Boyutlandırma ve Kontrast)
        image = image.resize((image.width * 2, image.height * 2), resample_filter)
        enhancer_sharp = ImageEnhance.Sharpness(image)
        image = enhancer_sharp.enhance(2.0)
        enhancer_cont = ImageEnhance.Contrast(image)
        image = enhancer_cont.enhance(1.5)
        
        try:
            text = pytesseract.image_to_string(image, lang='tur+eng')
        except pytesseract.TesseractError:
            text = pytesseract.image_to_string(image)
            
        # Tarih Bulma
        tarih_match = re.search(r'(\d{1,2})\s*[./-]\s*(\d{1,2})\s*[./-]\s*(\d{2,4})', text)
        if tarih_match:
            day, month, year = tarih_match.groups()
            if len(year) == 2: year = "20" + year
            tarih = f"{day.zfill(2)}.{month.zfill(2)}.{year}"
        else:
            tarih_match_2 = re.search(r'(\d{4})\s*[./-]\s*(\d{1,2})\s*[./-]\s*(\d{1,2})', text)
            if tarih_match_2:
                year, month, day = tarih_match_2.groups()
                tarih = f"{day.zfill(2)}.{month.zfill(2)}.{year}"
            
        # Tutar Bulma
        all_amounts = []
        matches = re.findall(r'(?:\b|[^0-9])(\d+(?:\s*[.,]\s*\d+)*)(?:\b|[^0-9])', text)
        for m in matches:
            val = parse_ocr_amount(m)
            if val is not None and val < 1000000:
                if 2000 <= val <= 2100 and val % 1 == 0:
                    continue
                if tarih and str(int(val)) in tarih:
                     if val % 1 == 0 and len(str(int(val))) == 4: continue
                all_amounts.append(val)

        lines = text.split('\n')
        found_amount = None
        
        for i, line in enumerate(lines):
            line_upper = line.upper()
            if any(k in line_upper for k in ["TOPLAM", "TUTAR", "TOP", "GENEL", "ODENECEK", "ODENEN", "KREDİ"]):
                line_matches = re.findall(r'(\d+(?:\s*[.,]\s*\d+)*)', line)
                valid_line_vals = []
                for m in line_matches:
                    val = parse_ocr_amount(m)
                    if val is not None and val < 1000000:
                        valid_line_vals.append(val)
                
                if valid_line_vals:
                    found_amount = max(valid_line_vals)
                    break
                
                if i + 1 < len(lines):
                    next_matches = re.findall(r'(\d+(?:\s*[.,]\s*\d+)*)', lines[i+1])
                    valid_next_vals = []
                    for m in next_matches:
                        val = parse_ocr_amount(m)
                        if val is not None and val < 1000000:
                            valid_next_vals.append(val)
                    
                    if valid_next_vals:
                        found_amount = max(valid_next_vals)
                        break
        
        if found_amount:
            tutar = found_amount
        elif all_amounts:
            tutar = max(all_amounts)
        
        # İşyeri/Açıklama Tahmini (İlk anlamlı satır)
        for line in lines:
            clean = line.strip()
            if clean and len(clean) > 3 and not any(x in clean.upper() for x in ["TARİH", "SAAT", "FİŞ", "NO", "TOPLAM"]):
                aciklama = clean[:50]
                break

    except Exception as e:
        print(f"OCR Hatası: {e}")
        # Hata olsa bile devam et, dosyayı kaydettik, kullanıcı düzenleyebilir.

    # Tarih formatı kontrolü (Geçersizse bugüne dön)
    try:
        datetime.strptime(tarih, "%d.%m.%Y")
    except:
        tarih = datetime.now().strftime("%d.%m.%Y")

    print(f"✅ Fiş Eklendi: {aciklama} - {tutar} TL - {tarih}")

    # 3. Veritabanına Ekle
    item_id = str(uuid.uuid4())
    c = conn.cursor()
    c.execute("INSERT INTO harcamalar (id, username, aciklama, tutar, kategori, tarih, fis_dosyasi) VALUES (?, ?, ?, ?, ?, ?, ?)",
              (item_id, username, aciklama, tutar, "Diğer", tarih, unique_filename))
    conn.commit()
    
    # Eklenen veriyi dön
    c.execute("SELECT * FROM harcamalar WHERE id = ?", (item_id,))
    row = dict(c.fetchone())
    return row

@app.get("/api/tahmin")
def butce_tahmini(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("SELECT * FROM harcamalar WHERE username = ?", (username,))
    harcamalar = [dict(row) for row in c.fetchall()]
    
    bugun = datetime.now()
    ay_sonu_gun = calendar.monthrange(bugun.year, bugun.month)[1]
    
    bu_ay_harcama = sum(h['tutar'] for h in harcamalar
                        if datetime.strptime(h['tarih'], "%d.%m.%Y").month == bugun.month and
                           datetime.strptime(h['tarih'], "%d.%m.%Y").year == bugun.year)

    gecen_gun = bugun.day
    gunluk_ort = bu_ay_harcama / gecen_gun if gecen_gun > 0 else 0
    tahmini_toplam = bu_ay_harcama + (gunluk_ort * (ay_sonu_gun - gecen_gun))
    
    return {
        "bu_ay_harcama": bu_ay_harcama,
        "gunluk_ortalama": gunluk_ort,
        "tahmini_ay_sonu": tahmini_toplam
    }

@app.get("/api/export/excel")
def export_excel(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("SELECT * FROM harcamalar WHERE username = ?", (username,))
    harcamalar = [dict(row) for row in c.fetchall()]
    c.execute("SELECT * FROM gelirler WHERE username = ?", (username,))
    gelirler = [dict(row) for row in c.fetchall()]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Harcamalar"
    
    # Başlıklar
    headers = ["Tarih", "Kategori", "Açıklama", "Tutar"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    
    for r, h in enumerate(harcamalar, 2):
        ws.cell(row=r, column=1, value=h.get("tarih"))
        ws.cell(row=r, column=2, value=h.get("kategori"))
        ws.cell(row=r, column=3, value=h.get("aciklama"))
        ws.cell(row=r, column=4, value=h.get("tutar"))
        
    # Gelirler Sayfası
    ws2 = wb.create_sheet("Gelirler")
    for col, h in enumerate(["Tarih", "Açıklama", "Tutar"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for r, g in enumerate(gelirler, 2):
        ws2.cell(row=r, column=1, value=g.get("tarih"))
        ws2.cell(row=r, column=2, value=g.get("aciklama"))
        ws2.cell(row=r, column=3, value=g.get("tutar"))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"butce_rapor_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/export/pdf")
def export_pdf(current_user: dict = Depends(get_current_user), conn: sqlite3.Connection = Depends(get_db)):
    username = current_user['username']
    c = conn.cursor()
    c.execute("SELECT * FROM harcamalar WHERE username = ?", (username,))
    harcamalar = [dict(row) for row in c.fetchall()]
    c.execute("SELECT * FROM gelirler WHERE username = ?", (username,))
    gelirler = [dict(row) for row in c.fetchall()]
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Başlık
    elements.append(Paragraph("Bütçe Raporu", styles['Title']))
    elements.append(Spacer(1, 12))
    
    # Harcamalar Tablosu
    elements.append(Paragraph("Harcamalar", styles['Heading2']))
    data = [["Tarih", "Kategori", "Açıklama", "Tutar"]]
    for h in harcamalar:
        data.append([h.get("tarih"), h.get("kategori"), h.get("aciklama"), str(h.get("tutar"))])
        
    if len(data) > 1:
        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("Harcama kaydı bulunamadı.", styles['Normal']))

    elements.append(Spacer(1, 12))

    # Gelirler Tablosu
    elements.append(Paragraph("Gelirler", styles['Heading2']))
    data_gelir = [["Tarih", "Açıklama", "Tutar"]]
    for g in gelirler:
        data_gelir.append([g.get("tarih"), g.get("aciklama"), str(g.get("tutar"))])

    if len(data_gelir) > 1:
        t2 = Table(data_gelir)
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(t2)
    else:
        elements.append(Paragraph("Gelir kaydı bulunamadı.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    
    filename = f"butce_rapor_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer, 
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Serve static PWA files from ./static
static_dir = os.path.join(BASE_DIR, "static")
if not os.path.exists(static_dir):
    os.makedirs(static_dir)

app.mount("/", StaticFiles(directory=static_dir, html=True), name="static")

if __name__ == "__main__":
    # Boş bir port bul
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.bind(('0.0.0.0', 0))
    port = sock.getsockname()[1]
    sock.close()

    # Bilgisayarın yerel IP adresini bulup ekrana yazdıralım
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        print(f"\n🚀 UYGULAMA HAZIR! Telefondan şu adrese gidin: http://{ip}:{port}\n")
    except Exception:
        print("\n⚠️ IP adresi otomatik bulunamadı. Lütfen 'ipconfig' ile kontrol edin.\n")

    # Sunucuyu tüm ağ arayüzlerine (0.0.0.0) açar, böylece telefondan erişilebilir.
    uvicorn.run(app, host="0.0.0.0", port=port)
