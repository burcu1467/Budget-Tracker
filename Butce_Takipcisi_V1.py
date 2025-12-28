# --- Bütçe Takipçisi V1.16 ---
# + CSV Export + Grafik/Chart + Tekrarlayan Harcamalar + Karşılaştırma + Bütçe Tahmini + Gelir Takibi + Kumbara + Excel Export

from datetime import datetime, timedelta
import json
import os
import csv
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

print("=" * 40)
print("Harcama Takip Programına Hoş Geldin!")
print("=" * 40)

harcamalar = []
gelirler = []
butce = 0
tekrarlayan_harcamalar = []
kumbara_bakiye = 0
kumbara_islemleri = []
kumbara_ayari = {"gunluk_tutar": 0, "haftalik_tutar": 0, "mod": None, "son_tarih": None}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOSYA_ADI = os.path.join(BASE_DIR, "harcamalar.json")
GELIR_DOSYASI = os.path.join(BASE_DIR, "gelirler.json")
BUTCE_DOSYASI = os.path.join(BASE_DIR, "butce.json")
TEKRARLAYAN_DOSYASI = os.path.join(BASE_DIR, "tekrarlayan.json")
KUMBARA_DOSYASI = os.path.join(BASE_DIR, "kumbara.json")

# Kullanılabilir kategoriler
KATEGORILER = ["Yemek", "Ulaşım", "Eğlence", "Alışveriş", "Sağlık", "Faturalar", "Diğer"]

def veri_yukle():
    """Dosyadan harcamaları, gelirleri, bütçeyi ve kumbara bilgilerini yükler"""
    global harcamalar, butce, gelirler, kumbara_bakiye, kumbara_islemleri, kumbara_ayari
    if os.path.exists(DOSYA_ADI):
        try:
            with open(DOSYA_ADI, "r", encoding="utf-8") as dosya:
                harcamalar = json.load(dosya)
                print(f"✓ {len(harcamalar)} harcama yüklendi.")
        except:
            print("⚠️ Dosya yüklenirken hata oluştu. Yeni başlıyorum.")
            harcamalar = []
    else:
        print("✓ Yeni bütçe takipçisi başlatılıyor.")
        harcamalar = []
    
    # Gelirleri yükle
    if os.path.exists(GELIR_DOSYASI):
        try:
            with open(GELIR_DOSYASI, "r", encoding="utf-8") as dosya:
                gelirler = json.load(dosya)
                print(f"✓ {len(gelirler)} gelir kaydı yüklendi.")
        except:
            gelirler = []
    else:
        gelirler = []
    
    # Bütçeyi yükle
    if os.path.exists(BUTCE_DOSYASI):
        try:
            with open(BUTCE_DOSYASI, "r", encoding="utf-8") as dosya:
                butce = json.load(dosya)
        except:
            butce = 0
    else:
        butce = 0
    
    # Tekrarlayan harcamaları yükle
    global tekrarlayan_harcamalar
    if os.path.exists(TEKRARLAYAN_DOSYASI):
        try:
            with open(TEKRARLAYAN_DOSYASI, "r", encoding="utf-8") as dosya:
                tekrarlayan_harcamalar = json.load(dosya)
        except:
            tekrarlayan_harcamalar = []
    else:
        tekrarlayan_harcamalar = []
    
    # Kumbara bilgilerini yükle
    if os.path.exists(KUMBARA_DOSYASI):
        try:
            with open(KUMBARA_DOSYASI, "r", encoding="utf-8") as dosya:
                kumbara_data = json.load(dosya)
                kumbara_bakiye = kumbara_data.get("bakiye", 0)
                kumbara_islemleri = kumbara_data.get("islemler", [])
                kumbara_ayari = kumbara_data.get("ayar", {"gunluk_tutar": 0, "haftalik_tutar": 0, "mod": None, "son_tarih": None})
        except:
            kumbara_bakiye = 0
            kumbara_islemleri = []
            kumbara_ayari = {"gunluk_tutar": 0, "haftalik_tutar": 0, "mod": None, "son_tarih": None}
    else:
        kumbara_bakiye = 0
        kumbara_islemleri = []
        kumbara_ayari = {"gunluk_tutar": 0, "haftalik_tutar": 0, "mod": None, "son_tarih": None}
    
    # Tekrarlayan harcamaları otomatik ekle
    otomatik_ekle_tekrarlayan()

def otomatik_ekle_tekrarlayan():
    """Bugünün tarihi tekrarlayan harcama günü ise otomatik ekler"""
    bugun = datetime.now()
    gun = bugun.day
    ay_gun = f"{gun}"
    
    for tekrarlayan in tekrarlayan_harcamalar:
        if not tekrarlayan['aktif']:
            continue
        
        if str(tekrarlayan['gun']) == ay_gun:
            # Bugün için zaten eklenmiş mi kontrol et
            bugune_ait = False
            bugune_git = bugun.strftime("%d.%m.%Y")
            
            for harcama in harcamalar:
                if (harcama['tarih'] == bugune_git and 
                    harcama['aciklama'] == tekrarlayan['aciklama'] and
                    harcama['kategori'] == tekrarlayan['kategori']):
                    bugune_ait = True
                    break
            
            if not bugune_ait:
                harcamalar.append({
                    "aciklama": tekrarlayan['aciklama'],
                    "tutar": tekrarlayan['tutar'],
                    "kategori": tekrarlayan['kategori'],
                    "tarih": bugune_git
                })
                veri_kaydet()

def veri_kaydet():
    """Harcamaları dosyaya kaydeder"""
    try:
        with open(DOSYA_ADI, "w", encoding="utf-8") as dosya:
            json.dump(harcamalar, dosya, ensure_ascii=False, indent=2)
    except:
        print("❌ Veriler kaydedilemedi!")

def gelir_kaydet():
    """Gelirleri dosyaya kaydeder"""
    try:
        with open(GELIR_DOSYASI, "w", encoding="utf-8") as dosya:
            json.dump(gelirler, dosya, ensure_ascii=False, indent=2)
    except:
        print("❌ Gelirler kaydedilemedi!")

def kumbara_kaydet():
    """Kumbara bilgilerini dosyaya kaydeder"""
    try:
        kumbara_data = {
            "bakiye": kumbara_bakiye,
            "islemler": kumbara_islemleri,
            "ayar": kumbara_ayari
        }
        with open(KUMBARA_DOSYASI, "w", encoding="utf-8") as dosya:
            json.dump(kumbara_data, dosya, ensure_ascii=False, indent=2)
    except:
        print("❌ Kumbara kaydedilemedi!")

def butce_kaydet():
    """Bütçeyi dosyaya kaydeder"""
    try:
        with open(BUTCE_DOSYASI, "w", encoding="utf-8") as dosya:
            json.dump(butce, dosya)
    except:
        print("❌ Bütçe kaydedilemedi!")

def tekrarlayan_kaydet():
    """Tekrarlayan harcamaları dosyaya kaydeder"""
    try:
        with open(TEKRARLAYAN_DOSYASI, "w", encoding="utf-8") as dosya:
            json.dump(tekrarlayan_harcamalar, dosya, ensure_ascii=False, indent=2)
    except:
        print("❌ Tekrarlayan harcamalar kaydedilemedi!")

def tekrarlayan_ekle():
    """Tekrarlayan harcama ekler"""
    print("\n--- TEKRARLAYAn HARCAMA EKLE ---")
    aciklama = input("Harcama adı (Örn: İnternet Faturası): ")
    
    kategorileri_goster()
    while True:
        try:
            kategori_idx = int(input("\nKategorisini seçiniz (1-7): ")) - 1
            if 0 <= kategori_idx < len(KATEGORILER):
                kategori = KATEGORILER[kategori_idx]
                break
            else:
                print("❌ Geçersiz kategori!")
        except ValueError:
            print("❌ Lütfen geçerli bir sayı giriniz!")
    
    while True:
        try:
            tutar = float(input("Aylık tutarı giriniz (TL): "))
            break
        except ValueError:
            print("❌ Lütfen geçerli bir sayı giriniz!")
    
    while True:
        try:
            gun = int(input("Ayın hangi günü eklenssin? (1-31): "))
            if 1 <= gun <= 31:
                break
            else:
                print("❌ Lütfen 1-31 arasında bir gün seçiniz!")
        except ValueError:
            print("❌ Lütfen geçerli bir sayı giriniz!")
    
    tekrarlayan_harcamalar.append({
        "aciklama": aciklama,
        "tutar": tutar,
        "kategori": kategori,
        "gun": gun,
        "aktif": True
    })
    tekrarlayan_kaydet()
    print(f"✓ '{aciklama}' tekrarlayan harcaması {gun}. günde eklenecek şekilde kaydedildi.")

def tekrarlayan_goster():
    """Tekrarlayan harcamaları gösterir"""
    if not tekrarlayan_harcamalar:
        print("\n❌ Tekrarlayan harcama kaydı yok.")
        return
    
    print("\n===== TEKRARLAYAn HARCAMALAR =====")
    for idx, h in enumerate(tekrarlayan_harcamalar, 1):
        status = "✓ Aktif" if h['aktif'] else "✗ İnaktif"
        print(f"{idx}. [{h['gun']}. gün] [{h['kategori']}] {h['aciklama']}: {h['tutar']} TL - {status}")

def tekrarlayan_duzenle():
    """Tekrarlayan harcamayı düzenler"""
    if not tekrarlayan_harcamalar:
        print("\n❌ Düzenlenecek tekrarlayan harcama yok.")
        return
    
    tekrarlayan_goster()
    try:
        idx = int(input("\nDüzenlenecek harcamanın numarasını giriniz: ")) - 1
        if 0 <= idx < len(tekrarlayan_harcamalar):
            h = tekrarlayan_harcamalar[idx]
            
            aciklama = input(f"Yeni adı (Eski: {h['aciklama']}): ").strip()
            if aciklama:
                h['aciklama'] = aciklama
            
            tutar_str = input(f"Yeni tutar (Eski: {h['tutar']}): ").strip()
            if tutar_str:
                h['tutar'] = float(tutar_str)
            
            gun_str = input(f"Yeni gün (Eski: {h['gun']}): ").strip()
            if gun_str:
                h['gun'] = int(gun_str)
            
            print("✓ Tekrarlayan harcama güncellendi.")
            tekrarlayan_kaydet()
        else:
            print("❌ Geçersiz numara!")
    except ValueError:
        print("❌ Lütfen geçerli bir sayı giriniz!")

def tekrarlayan_sil():
    """Tekrarlayan harcamayı siler"""
    if not tekrarlayan_harcamalar:
        print("\n❌ Silinecek tekrarlayan harcama yok.")
        return
    
    tekrarlayan_goster()
    try:
        idx = int(input("\nSilinecek harcamanın numarasını giriniz: ")) - 1
        if 0 <= idx < len(tekrarlayan_harcamalar):
            silinen = tekrarlayan_harcamalar.pop(idx)
            print(f"✓ '{silinen['aciklama']}' tekrarlayan harcaması silindi.")
            tekrarlayan_kaydet()
        else:
            print("❌ Geçersiz numara!")
    except ValueError:
        print("❌ Lütfen geçerli bir sayı giriniz!")

def tekrarlayan_menu():
    """Tekrarlayan harcama yönetim menüsü"""
    while True:
        print("\n" + "=" * 40)
        print("1. Tekrarlayan Harcama Ekle")
        print("2. Tekrarlayan Harcamaları Gör")
        print("3. Tekrarlayan Harcama Düzenle")
        print("4. Tekrarlayan Harcama Sil")
        print("5. Geri Dön")
        print("=" * 40)
        secim = input("Seçiminiz: ")
        
        if secim == "1":
            tekrarlayan_ekle()
        elif secim == "2":
            tekrarlayan_goster()
        elif secim == "3":
            tekrarlayan_duzenle()
        elif secim == "4":
            tekrarlayan_sil()
        elif secim == "5":
            break
        else:
            print("❌ Geçersiz seçim!")

def butce_belirle():
    """Bütçe belirler"""
    global butce
    print("\n--- BÜTÇE BELIRLE ---")
    try:
        yeni_butce = float(input("Aylık bütçeyi giriniz (TL): "))
        if yeni_butce >= 0:
            butce = yeni_butce
            butce_kaydet()
            print(f"✓ Bütçe {yeni_butce} TL olarak belirlendi.")
        else:
            print("❌ Bütçe negatif olamaz!")
    except ValueError:
        print("❌ Lütfen geçerli bir sayı giriniz!")

def butce_goster():
    """Bütçe bilgilerini gösterir"""
    if butce == 0:
        print("\n❌ Henüz bütçe belirlenmedi.")
        return
    
    genel_toplam = sum(h['tutar'] for h in harcamalar)
    kalan_butce = butce - genel_toplam
    yuzde = (genel_toplam / butce * 100) if butce > 0 else 0
    
    print("\n===== BÜTÇE DURUMU =====")
    print(f"📊 Belirlenen Bütçe: {butce} TL")
    print(f"💸 Harcanan Toplam: {genel_toplam} TL")
    print(f"💰 Kalan Bütçe: {kalan_butce} TL")
    print(f"📈 Kullanılan Oran: %{yuzde:.1f}")
    
    if kalan_butce < 0:
        print(f"⚠️ BÜTÇESİ AŞTINIZ! {abs(kalan_butce)} TL FAZLA HARCAMA!")
    elif kalan_butce < butce * 0.2:
        print("⚠️ DİKKAT: Bütçenizin %80'ini kullandınız!")
    else:
        print("✓ Bütçenizde yer var.")

def kategorileri_goster():
    """Kullanılabilir kategorileri gösterir"""
    print("\n📂 Kullanılabilir Kategoriler:")
    for idx, kategori in enumerate(KATEGORILER, 1):
        print(f"   {idx}. {kategori}")

def harcama_ekle():
    """Yeni harcama ekler"""
    print("\n--- HARCAMA EKLE ---")
    aciklama = input("Harcama yerini giriniz (Örn: Şirnak Marketim): ")
    
    kategorileri_goster()
    while True:
        try:
            kategori_idx = int(input("\nKategorisini seçiniz (1-7): ")) - 1
            if 0 <= kategori_idx < len(KATEGORILER):
                kategori = KATEGORILER[kategori_idx]
                break
            else:
                print("❌ Geçersiz kategori! Lütfen 1-7 arasında seçim yapınız.")
        except ValueError:
            print("❌ Lütfen geçerli bir sayı giriniz!")
    
    while True:
        try:
            tutar = float(input("Harcama tutarını giriniz (Örn: 50): "))
            break
        except ValueError:
            print("Lütfen geçerli bir sayı giriniz!")
    
    # Bütçe kontrolü
    if butce > 0:
        harcanan = sum(h['tutar'] for h in harcamalar)
        genel_toplam = harcanan + tutar
        kalan = butce - harcanan
        yuzde_sonra = (genel_toplam / butce * 100) if butce > 0 else 0
        
        if genel_toplam > butce:
            print(f"⚠️ UYARI: Bu harcamayı yaparsanız bütçeyi aşacaksınız!")
            print(f"   Kalan bütçe: {kalan} TL, Harcama: {tutar} TL")
            devam = input("Yine de eklemek istiyor musunuz? (E/H): ").strip().lower()
            if devam not in ["e", "evet"]:
                print("❌ Harcama eklenmedi.")
                return
        elif yuzde_sonra >= 80:
            print(f"⚠️ DİKKAT: Bütçenizin %80'ini aşacaksınız!")
            print(f"   Kullanılan oran: %{yuzde_sonra:.1f}")
    
    # Tarih seçeneği
    tarih_secim = input("Harcama tarihini giriniz (Boş bırakırsa bugün: GG.AA.YYYY veya Enter): ").strip()
    
    if tarih_secim == "":
        tarih = datetime.now().strftime("%d.%m.%Y")
    else:
        try:
            # Tarih formatını kontrol et
            datetime.strptime(tarih_secim, "%d.%m.%Y")
            tarih = tarih_secim
        except ValueError:
            print("❌ Geçersiz tarih! Bugünün tarihini kullanıyorum.")
            tarih = datetime.now().strftime("%d.%m.%Y")
    
    harcamalar.append({"aciklama": aciklama, "tutar": tutar, "kategori": kategori, "tarih": tarih})
    print(f"✓ [{tarih}] [{kategori}] {aciklama} için {tutar} TL harcama kaydedildi.")
    veri_kaydet()

def gelir_ekle():
    """Yeni gelir ekler"""
    print("\n--- GELİR EKLE ---")
    aciklama = input("Gelir kaynağını giriniz (Örn: Maaş, Ek iş, Satış): ")
    
    while True:
        try:
            tutar = float(input("Gelir tutarını giriniz (Örn: 5000): "))
            break
        except ValueError:
            print("❌ Lütfen geçerli bir sayı giriniz!")
    
    # Tarih seçeneği
    tarih_secim = input("Gelir tarihini giriniz (Boş bırakırsa bugün: GG.AA.YYYY veya Enter): ").strip()
    
    if tarih_secim == "":
        tarih = datetime.now().strftime("%d.%m.%Y")
    else:
        try:
            # Tarih formatını kontrol et
            datetime.strptime(tarih_secim, "%d.%m.%Y")
            tarih = tarih_secim
        except ValueError:
            print("❌ Geçersiz tarih! Bugünün tarihini kullanıyorum.")
            tarih = datetime.now().strftime("%d.%m.%Y")
    
    gelirler.append({"aciklama": aciklama, "tutar": tutar, "tarih": tarih})
    print(f"✓ [{tarih}] {aciklama} için {tutar} TL gelir kaydedildi.")
    gelir_kaydet()

def gelir_goster():
    """Tüm gelirleri gösterir"""
    if not gelirler:
        print("\n❌ Henüz gelir kaydı yok.")
        return
    
    print("\n===== TÜM GELİRLER =====")
    toplam = 0
    for idx, gelir in enumerate(gelirler, 1):
        print(f"{idx}. [{gelir['tarih']}] {gelir['aciklama']}: {gelir['tutar']} TL")
        toplam += gelir['tutar']
    
    print(f"\n💰 TOPLAM GELİR: {toplam:.2f} TL")

def gider_gelir_ozeti():
    """Gider vs Gelir özeti gösterir"""
    toplam_gider = sum(h['tutar'] for h in harcamalar)
    toplam_gelir = sum(g['tutar'] for g in gelirler)
    fark = toplam_gelir - toplam_gider
    
    print("\n" + "=" * 50)
    print("💼 GİDER VE GELİR ÖZETİ")
    print("=" * 50)
    print(f"📊 Toplam Gelir:        {toplam_gelir:>15.2f} TL")
    print(f"💸 Toplam Gider:        {toplam_gider:>15.2f} TL")
    print("-" * 50)
    
    if fark >= 0:
        print(f"✅ NET KAZANÇ:         {fark:>15.2f} TL")
        if toplam_gelir > 0:
            oran = (toplam_gider / toplam_gelir) * 100
            print(f"📈 Gider Oranı:        {oran:>14.1f}%")
    else:
        print(f"❌ NET ZARAR:          {fark:>15.2f} TL")
        if toplam_gelir > 0:
            oran = (toplam_gider / toplam_gelir) * 100
            print(f"📉 Gider Oranı:        {oran:>14.1f}%")
    
    print("=" * 50)
    
    # İstatistik
    if gelirler and harcamalar:
        ortalama_gelir = toplam_gelir / len(gelirler)
        ortalama_gider = toplam_gider / len(harcamalar)
        print(f"\n📊 İstatistikler:")
        print(f"   Gelir Kaynağı Sayısı: {len(gelirler)}")
        print(f"   Ortalama Gelir: {ortalama_gelir:.2f} TL")
        print(f"   Harcama Sayısı: {len(harcamalar)}")
        print(f"   Ortalama Harcama: {ortalama_gider:.2f} TL")

def kumbara_ayarla():
    """Kumbara tasarruf ayarlarını belirler"""
    global kumbara_ayari, kumbara_bakiye
    
    print("\n" + "=" * 50)
    print("🏺 KUMBARA AYARLARI")
    print("=" * 50)
    print("1. Günlük Tasarruf Modu")
    print("2. Haftalık Tasarruf Modu")
    print("3. Geri Dön")
    
    secim = input("Seçiminiz: ")
    
    if secim == "1":
        try:
            tutar = float(input("Günde ne kadar tasarruf etmek istiyorsunuz? (TL): "))
            if tutar > 0:
                kumbara_ayari["gunluk_tutar"] = tutar
                kumbara_ayari["haftalik_tutar"] = 0
                kumbara_ayari["mod"] = "gunluk"
                kumbara_ayari["son_tarih"] = datetime.now().strftime("%d.%m.%Y")
                kumbara_kaydet()
                print(f"✓ Günlük {tutar} TL tasarruf modu aktifleştirildi.")
            else:
                print("❌ Tutar 0'dan büyük olmalı!")
        except ValueError:
            print("❌ Lütfen geçerli bir sayı giriniz!")
    
    elif secim == "2":
        try:
            tutar = float(input("Haftada ne kadar tasarruf etmek istiyorsunuz? (TL): "))
            if tutar > 0:
                kumbara_ayari["haftalik_tutar"] = tutar
                kumbara_ayari["gunluk_tutar"] = 0
                kumbara_ayari["mod"] = "haftalik"
                kumbara_ayari["son_tarih"] = datetime.now().strftime("%d.%m.%Y")
                kumbara_kaydet()
                print(f"✓ Haftalık {tutar} TL tasarruf modu aktifleştirildi.")
            else:
                print("❌ Tutar 0'dan büyük olmalı!")
        except ValueError:
            print("❌ Lütfen geçerli bir sayı giriniz!")
    
    elif secim == "3":
        pass
    else:
        print("❌ Geçersiz seçim!")

def kumbara_ekle():
    """Kumbaraya otomatik olarak para ekler"""
    global kumbara_bakiye, kumbara_islemleri, kumbara_ayari
    
    if kumbara_ayari["mod"] is None:
        print("❌ Lütfen önce kumbara ayarlarını belirleyiniz!")
        return
    
    son_tarih = datetime.strptime(kumbara_ayari["son_tarih"], "%d.%m.%Y") if kumbara_ayari["son_tarih"] else None
    bugün = datetime.now()
    
    tutar_eklenecek = 0
    
    if kumbara_ayari["mod"] == "gunluk":
        # Son para eklenme tarihinden günümüze kadar kaç gün geçti?
        if son_tarih:
            gün_farkı = (bugün - son_tarih).days
        else:
            gün_farkı = 1
        
        if gün_farkı > 0:
            tutar_eklenecek = kumbara_ayari["gunluk_tutar"] * gün_farkı
    
    elif kumbara_ayari["mod"] == "haftalik":
        # Son para eklenme tarihinden günümüze kadar kaç hafta geçti?
        if son_tarih:
            hafta_farkı = (bugün - son_tarih).days // 7
        else:
            hafta_farkı = 1
        
        if hafta_farkı > 0:
            tutar_eklenecek = kumbara_ayari["haftalik_tutar"] * hafta_farkı
    
    if tutar_eklenecek > 0:
        kumbara_bakiye += tutar_eklenecek
        kumbara_islemleri.append({
            "tarih": bugün.strftime("%d.%m.%Y"),
            "tutar": tutar_eklenecek,
            "tür": "ekleme",
            "açıklama": "Otomatik tasarruf"
        })
        kumbara_ayari["son_tarih"] = bugün.strftime("%d.%m.%Y")
        kumbara_kaydet()
        print(f"✓ {tutar_eklenecek:.2f} TL kumbaraya eklendi!")

def kumbara_goster():
    """Kumbara bilgilerini gösterir"""
    print("\n" + "=" * 50)
    print("🏺 KUMBARA DURUMU")
    print("=" * 50)
    print(f"💰 Kumbara Bakiyesi: {kumbara_bakiye:.2f} TL")
    
    if kumbara_ayari["mod"]:
        print(f"📌 Tasarruf Modu: {('Günlük' if kumbara_ayari['mod'] == 'gunluk' else 'Haftalık')}")
        if kumbara_ayari["mod"] == "gunluk":
            print(f"📊 Günlük Tutar: {kumbara_ayari['gunluk_tutar']:.2f} TL")
        else:
            print(f"📊 Haftalık Tutar: {kumbara_ayari['haftalik_tutar']:.2f} TL")
        print(f"📅 Son Güncelleme: {kumbara_ayari['son_tarih']}")
    else:
        print("⚠️ Henüz tasarruf modu belirlenmiş değil!")
    
    if kumbara_islemleri:
        print(f"\n📝 Son 10 İşlem:")
        for işlem in kumbara_islemleri[-10:]:
            tür = "➕" if işlem["tür"] == "ekleme" else "➖"
            print(f"{tür} [{işlem['tarih']}] {işlem['açıklama']}: {işlem['tutar']:.2f} TL")
    
    print("=" * 50)

def kumbara_para_cek():
    """Kumbaradan para çekmek için"""
    global kumbara_bakiye, kumbara_islemleri
    
    print(f"\n🏺 Kumbara Bakiyesi: {kumbara_bakiye:.2f} TL")
    try:
        tutar = float(input("Çekmek istediğiniz tutarı giriniz (TL): "))
        if tutar <= 0:
            print("❌ Tutar 0'dan büyük olmalı!")
            return
        
        if tutar > kumbara_bakiye:
            print(f"❌ Yetersiz bakiye! Bakiye: {kumbara_bakiye:.2f} TL")
            return
        
        kumbara_bakiye -= tutar
        kumbara_islemleri.append({
            "tarih": datetime.now().strftime("%d.%m.%Y"),
            "tutar": tutar,
            "tür": "çekme",
            "açıklama": "Para çekildi"
        })
        kumbara_kaydet()
        print(f"✓ {tutar:.2f} TL kumbaradan çekildi!")
        print(f"📊 Yeni Bakiye: {kumbara_bakiye:.2f} TL")
    except ValueError:
        print("❌ Lütfen geçerli bir sayı giriniz!")

def kumbara_menu():
    """Kumbara yönetim menüsü"""
    while True:
        print("\n" + "=" * 40)
        print("1. Kumbara Ayarlarını Belirle")
        print("2. Kumbara Durumunu Gör")
        print("3. Para Çek")
        print("4. Geri Dön")
        print("=" * 40)
        secim = input("Seçiminiz: ")
        
        if secim == "1":
            kumbara_ayarla()
        elif secim == "2":
            kumbara_goster()
        elif secim == "3":
            kumbara_para_cek()
        elif secim == "4":
            break
        else:
            print("❌ Geçersiz seçim!")

def harcamaları_goster():
    """Tüm harcamaları gösterir"""
    if not harcamalar:
        print("\n❌ Henüz harcama kaydı yok.")
        return
    
    print("\n===== TÜM HARCAMALAR =====")
    toplam = 0
    for idx, harcama in enumerate(harcamalar, 1):
        print(f"{idx}. [{harcama['tarih']}] [{harcama['kategori']}] {harcama['aciklama']}: {harcama['tutar']} TL")
        toplam += harcama['tutar']
    
    print(f"\n💰 Toplam Harcama: {toplam} TL")

def kategori_bazli_goster():
    """Kategori bazlı harcamaları gösterir"""
    if not harcamalar:
        print("\n❌ Henüz harcama kaydı yok.")
        return
    
    kategori_toplamı = {}
    for harcama in harcamalar:
        kat = harcama['kategori']
        if kat not in kategori_toplamı:
            kategori_toplamı[kat] = 0
        kategori_toplamı[kat] += harcama['tutar']
    
    print("\n===== KATEGORİ BAZLI HARCAMALAR =====")
    genel_toplam = 0
    for kategori, toplam in sorted(kategori_toplamı.items()):
        print(f"📂 {kategori}: {toplam} TL")
        genel_toplam += toplam
    
    print(f"\n💰 Genel Toplam: {genel_toplam} TL")

def istatistikler_goster():
    """İstatistikleri gösterir"""
    if not harcamalar:
        print("\n❌ Henüz harcama kaydı yok.")
        return
    
    # Kategori bazlı istatistikler
    kategori_toplamı = {}
    kategori_sayısı = {}
    
    for harcama in harcamalar:
        kat = harcama['kategori']
        if kat not in kategori_toplamı:
            kategori_toplamı[kat] = 0
            kategori_sayısı[kat] = 0
        kategori_toplamı[kat] += harcama['tutar']
        kategori_sayısı[kat] += 1
    
    print("\n===== İSTATİSTİKLER =====")
    
    # Genel toplam
    genel_toplam = sum(h['tutar'] for h in harcamalar)
    print(f"\n💰 Genel Toplam: {genel_toplam} TL")
    print(f"📊 Toplam Harcama Sayısı: {len(harcamalar)}")
    print(f"📈 Ortalama Harcama: {genel_toplam / len(harcamalar):.2f} TL")
    
    # Kategori bazlı ortalamalar
    print("\n📂 KATEGORİ ORTALAMALARI:")
    for kategori in sorted(kategori_toplamı.keys()):
        toplam = kategori_toplamı[kategori]
        sayı = kategori_sayısı[kategori]
        ortalama = toplam / sayı
        print(f"   {kategori}: {ortalama:.2f} TL (Toplam: {toplam} TL, {sayı} işlem)")
    
    # En çok harcanan kategori
    max_kategori = max(kategori_toplamı, key=kategori_toplamı.get)
    print(f"\n🔝 En Çok Harcanan: {max_kategori} ({kategori_toplamı[max_kategori]} TL)")
    
    # En az harcanan kategori
    min_kategori = min(kategori_toplamı, key=kategori_toplamı.get)
    print(f"🔻 En Az Harcanan: {min_kategori} ({kategori_toplamı[min_kategori]} TL)")
    
    # En yüksek tek harcama
    max_harcama = max(harcamalar, key=lambda x: x['tutar'])
    print(f"\n💸 En Yüksek Harcama: [{max_harcama['tarih']}] {max_harcama['aciklama']} ({max_harcama['tutar']} TL)")
    
    # En düşük tek harcama
    min_harcama = min(harcamalar, key=lambda x: x['tutar'])
    print(f"💵 En Düşük Harcama: [{min_harcama['tarih']}] {min_harcama['aciklama']} ({min_harcama['tutar']} TL)")

def harcama_sil():
    """Harcama siler"""
    if not harcamalar:
        print("\n❌ Silinecek harcama yok.")
        return
    
    harcamaları_goster()
    try:
        idx = int(input("\nSilinecek harcamanın numarasını giriniz: ")) - 1
        if 0 <= idx < len(harcamalar):
            silinen = harcamalar.pop(idx)
            print(f"✓ '{silinen['aciklama']}' harcaması silindi.")
            veri_kaydet()
        else:
            print("❌ Geçersiz numara!")
    except ValueError:
        print("❌ Lütfen geçerli bir sayı giriniz!")

def harcama_duzenle():
    """Harcamayı düzenler"""
    if not harcamalar:
        print("\n❌ Düzenlenecek harcama yok.")
        return
    
    harcamaları_goster()
    try:
        idx = int(input("\nDüzenlenecek harcamanın numarasını giriniz: ")) - 1
        if 0 <= idx < len(harcamalar):
            aciklama = input("Yeni harcama yerini giriniz (boş bırakırsa eski kalır): ")
            if aciklama:
                harcamalar[idx]["aciklama"] = aciklama
            
            kategorileri_goster()
            kategori_str = input("\nYeni kategoriyi seçiniz (boş bırakırsa eski kalır): ")
            if kategori_str:
                try:
                    kategori_idx = int(kategori_str) - 1
                    if 0 <= kategori_idx < len(KATEGORILER):
                        harcamalar[idx]["kategori"] = KATEGORILER[kategori_idx]
                except ValueError:
                    pass
            
            tutar_str = input("Yeni tutarı giriniz (boş bırakırsa eski kalır): ")
            if tutar_str:
                harcamalar[idx]["tutar"] = float(tutar_str)
            
            tarih_str = input("Yeni tarihi giriniz (boş bırakırsa eski kalır - GG.AA.YYYY): ")
            if tarih_str:
                try:
                    datetime.strptime(tarih_str, "%d.%m.%Y")
                    harcamalar[idx]["tarih"] = tarih_str
                except ValueError:
                    print("❌ Geçersiz tarih formatı! Eski tarih korundu.")
            
            print("✓ Harcama güncellendi.")
            veri_kaydet()
        else:
            print("❌ Geçersiz numara!")
    except ValueError:
        print("❌ Lütfen geçerli bir sayı giriniz!")

def tarih_araligi_goster():
    """Belirli tarih aralığında harcamaları gösterir"""
    if not harcamalar:
        print("\n❌ Henüz harcama kaydı yok.")
        return
    
    print("\n--- TARİH ARALIGI FİLTRESİ ---")
    
    while True:
        basla_tarih_str = input("Başlangıç tarihini giriniz (GG.AA.YYYY): ").strip()
        try:
            basla_tarih = datetime.strptime(basla_tarih_str, "%d.%m.%Y")
            break
        except ValueError:
            print("❌ Geçersiz tarih formatı! Örnek: 01.12.2025")
    
    while True:
        bitis_tarih_str = input("Bitiş tarihini giriniz (GG.AA.YYYY): ").strip()
        try:
            bitis_tarih = datetime.strptime(bitis_tarih_str, "%d.%m.%Y")
            break
        except ValueError:
            print("❌ Geçersiz tarih formatı! Örnek: 31.12.2025")
    
    # Tarih aralığında harcamaları filtrele
    filtrelenen = []
    for harcama in harcamalar:
        h_tarih = datetime.strptime(harcama['tarih'], "%d.%m.%Y")
        if basla_tarih <= h_tarih <= bitis_tarih:
            filtrelenen.append(harcama)
    
    if not filtrelenen:
        print(f"\n❌ {basla_tarih_str} - {bitis_tarih_str} aralığında harcama bulunamadı.")
        return
    
    print(f"\n===== {basla_tarih_str} - {bitis_tarih_str} ARALIĞI HARCAMALAR =====")
    toplam = 0
    for idx, harcama in enumerate(filtrelenen, 1):
        print(f"{idx}. [{harcama['tarih']}] [{harcama['kategori']}] {harcama['aciklama']}: {harcama['tutar']} TL")
        toplam += harcama['tutar']
    
    print(f"\n💰 Aralık Toplam Harcama: {toplam} TL")
    
    # Kategori analizi
    kategori_toplamı = {}
    for harcama in filtrelenen:
        kat = harcama['kategori']
        if kat not in kategori_toplamı:
            kategori_toplamı[kat] = 0
        kategori_toplamı[kat] += harcama['tutar']
    
    print("\n📂 Kategori Dağılımı:")
    for kategori in sorted(kategori_toplamı.keys()):
        print(f"   {kategori}: {kategori_toplamı[kategori]} TL")

def csv_export():
    """Harcamaları CSV dosyasına aktarır"""
    if not harcamalar:
        print("\n❌ Aktarılacak harcama yok.")
        return
    
    print("\n--- CSV EXPORT ---")
    print("1. Tüm harcamaları export et")
    print("2. Tarih aralığına göre export et")
    secim = input("Seçiminiz (1-2): ")
    
    if secim == "1":
        veri_export = harcamalar
        dosya_adi = f"harcamalar_{datetime.now().strftime('%d_%m_%Y')}.csv"
    elif secim == "2":
        while True:
            basla_tarih_str = input("Başlangıç tarihini giriniz (GG.AA.YYYY): ").strip()
            try:
                basla_tarih = datetime.strptime(basla_tarih_str, "%d.%m.%Y")
                break
            except ValueError:
                print("❌ Geçersiz tarih formatı! Örnek: 01.12.2025")
        
        while True:
            bitis_tarih_str = input("Bitiş tarihini giriniz (GG.AA.YYYY): ").strip()
            try:
                bitis_tarih = datetime.strptime(bitis_tarih_str, "%d.%m.%Y")
                break
            except ValueError:
                print("❌ Geçersiz tarih formatı! Örnek: 31.12.2025")
        
        veri_export = []
        for harcama in harcamalar:
            h_tarih = datetime.strptime(harcama['tarih'], "%d.%m.%Y")
            if basla_tarih <= h_tarih <= bitis_tarih:
                veri_export.append(harcama)
        
        if not veri_export:
            print(f"❌ {basla_tarih_str} - {bitis_tarih_str} aralığında harcama bulunamadı.")
            return
        
        dosya_adi = f"harcamalar_{basla_tarih_str.replace('.', '_')}_to_{bitis_tarih_str.replace('.', '_')}.csv"
    else:
        print("❌ Geçersiz seçim!")
        return
    
    try:
        with open(dosya_adi, "w", newline="", encoding="utf-8-sig") as dosya:
            yazici = csv.writer(dosya, delimiter=",")
            
            # Başlık satırı
            yazici.writerow(["Tarih", "Kategori", "Açıklama", "Tutar (TL)"])
            
            # Veri satırları
            for harcama in veri_export:
                yazici.writerow([
                    harcama['tarih'],
                    harcama['kategori'],
                    harcama['aciklama'],
                    harcama['tutar']
                ])
            
            # Özet satırı
            toplam = sum(h['tutar'] for h in veri_export)
            yazici.writerow([])
            yazici.writerow(["TOPLAM", "", "", toplam])
        
        print(f"✓ Veriler '{dosya_adi}' dosyasına başarıyla aktarıldı!")
        print(f"  {len(veri_export)} harcama, Toplam: {toplam} TL")
    except Exception as e:
        print(f"❌ Export sırasında hata oluştu: {e}")

def excel_export():
    """Tüm verileri Excel dosyasına aktarır"""
    print("\n--- EXCEL'E AKTAR ---")
    print("1. Tüm Veriler")
    print("2. Tarih Aralığını Filtrele")
    secim = input("Seçiminiz: ")
    
    veri_export = []
    
    if secim == "1":
        veri_export = harcamalar
    elif secim == "2":
        basla_tarih_str = input("Başlangıç tarihi giriniz (GG.AA.YYYY): ")
        bitis_tarih_str = input("Bitiş tarihi giriniz (GG.AA.YYYY): ")
        
        try:
            basla_tarih = datetime.strptime(basla_tarih_str, "%d.%m.%Y")
            bitis_tarih = datetime.strptime(bitis_tarih_str, "%d.%m.%Y")
            
            for harcama in harcamalar:
                h_tarih = datetime.strptime(harcama['tarih'], "%d.%m.%Y")
                if basla_tarih <= h_tarih <= bitis_tarih:
                    veri_export.append(harcama)
        except ValueError:
            print("❌ Geçersiz tarih formatı!")
            return
    else:
        print("❌ Geçersiz seçim!")
        return
    
    if not veri_export:
        print("❌ Aktarılacak veri yok!")
        return
    
    try:
        # Excel Workbook oluştur
        wb = Workbook()
        
        # ===== HARCAMALAR SAYFASI =====
        ws_harcama = wb.active
        ws_harcama.title = "Harcamalar"
        
        # Başlık formatı
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Başlıklar
        headers = ["Tarih", "Kategori", "Açıklama", "Tutar (TL)"]
        for col, header in enumerate(headers, 1):
            cell = ws_harcama.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Veri satırları
        for row, harcama in enumerate(veri_export, 2):
            ws_harcama.cell(row=row, column=1, value=harcama['tarih'])
            ws_harcama.cell(row=row, column=2, value=harcama['kategori'])
            ws_harcama.cell(row=row, column=3, value=harcama['aciklama'])
            ws_harcama.cell(row=row, column=4, value=harcama['tutar'])
            
            for col in range(1, 5):
                ws_harcama.cell(row=row, column=col).border = border
        
        # Toplam satırı
        toplam_row = len(veri_export) + 2
        ws_harcama.cell(row=toplam_row, column=1, value="TOPLAM")
        toplam = sum(h['tutar'] for h in veri_export)
        ws_harcama.cell(row=toplam_row, column=4, value=toplam)
        
        for col in range(1, 5):
            cell = ws_harcama.cell(row=toplam_row, column=col)
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(bold=True, size=11)
            cell.border = border
        
        # Sütun genişliklerini ayarla
        ws_harcama.column_dimensions['A'].width = 12
        ws_harcama.column_dimensions['B'].width = 15
        ws_harcama.column_dimensions['C'].width = 25
        ws_harcama.column_dimensions['D'].width = 12
        
        # ===== GELİRLER SAYFASI =====
        ws_gelir = wb.create_sheet("Gelirler")
        
        headers_gelir = ["Tarih", "Kaynağı", "Tutar (TL)"]
        for col, header in enumerate(headers_gelir, 1):
            cell = ws_gelir.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        for row, gelir in enumerate(gelirler, 2):
            ws_gelir.cell(row=row, column=1, value=gelir['tarih'])
            ws_gelir.cell(row=row, column=2, value=gelir['aciklama'])
            ws_gelir.cell(row=row, column=3, value=gelir['tutar'])
            
            for col in range(1, 4):
                ws_gelir.cell(row=row, column=col).border = border
        
        # Gelir toplam satırı
        if gelirler:
            gelir_toplam_row = len(gelirler) + 2
            ws_gelir.cell(row=gelir_toplam_row, column=1, value="TOPLAM")
            gelir_toplam = sum(g['tutar'] for g in gelirler)
            ws_gelir.cell(row=gelir_toplam_row, column=3, value=gelir_toplam)
            
            for col in range(1, 4):
                cell = ws_gelir.cell(row=gelir_toplam_row, column=col)
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(bold=True, size=11)
                cell.border = border
        
        ws_gelir.column_dimensions['A'].width = 12
        ws_gelir.column_dimensions['B'].width = 20
        ws_gelir.column_dimensions['C'].width = 12
        
        # ===== KUMBARA SAYFASI =====
        ws_kumbara = wb.create_sheet("Kumbara")
        
        ws_kumbara.cell(row=1, column=1, value="Kumbara Bakiyesi").font = Font(bold=True, size=12)
        ws_kumbara.cell(row=1, column=2, value=kumbara_bakiye).font = Font(size=12)
        
        ws_kumbara.cell(row=3, column=1, value="Tasarruf Modu").font = Font(bold=True)
        mod_text = f"Günlük: {kumbara_ayari['gunluk_tutar']} TL" if kumbara_ayari['mod'] == 'gunluk' else f"Haftalık: {kumbara_ayari['haftalik_tutar']} TL"
        ws_kumbara.cell(row=3, column=2, value=mod_text)
        
        # Kumbara işlemleri
        if kumbara_islemleri:
            ws_kumbara.cell(row=5, column=1, value="İşlem Tarihi").font = Font(bold=True)
            ws_kumbara.cell(row=5, column=2, value="Tür").font = Font(bold=True)
            ws_kumbara.cell(row=5, column=3, value="Tutar").font = Font(bold=True)
            
            for row, islem in enumerate(kumbara_islemleri, 6):
                ws_kumbara.cell(row=row, column=1, value=islem['tarih'])
                ws_kumbara.cell(row=row, column=2, value=islem['tür'])
                ws_kumbara.cell(row=row, column=3, value=islem['tutar'])
        
        ws_kumbara.column_dimensions['A'].width = 15
        ws_kumbara.column_dimensions['B'].width = 20
        ws_kumbara.column_dimensions['C'].width = 12

        # ===== TEKRARLAYAN SAYFASI =====
        ws_tekr = wb.create_sheet("Tekrarlayan")
        headers_tekr = ["Gün", "Kategori", "Açıklama", "Tutar", "Aktif"]
        for col, header in enumerate(headers_tekr, 1):
            cell = ws_tekr.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align

        for row, t in enumerate(tekrarlayan_harcamalar, 2):
            ws_tekr.cell(row=row, column=1, value=t.get('gun'))
            ws_tekr.cell(row=row, column=2, value=t.get('kategori'))
            ws_tekr.cell(row=row, column=3, value=t.get('aciklama'))
            ws_tekr.cell(row=row, column=4, value=t.get('tutar'))
            ws_tekr.cell(row=row, column=5, value=('Evet' if t.get('aktif') else 'Hayır'))
            for col in range(1, 6):
                ws_tekr.cell(row=row, column=col).border = border

        ws_tekr.column_dimensions['A'].width = 8
        ws_tekr.column_dimensions['B'].width = 15
        ws_tekr.column_dimensions['C'].width = 30
        ws_tekr.column_dimensions['D'].width = 12
        ws_tekr.column_dimensions['E'].width = 8

        # ===== KATEGORI ÖZETİ =====
        ws_kat = wb.create_sheet("Kategori Özeti")
        ws_kat.cell(row=1, column=1, value="Kategori").font = header_font
        ws_kat.cell(row=1, column=2, value="Toplam Harcama").font = header_font
        kategori_toplam = {}
        for h in veri_export:
            kat = h.get('kategori', 'Diğer')
            kategori_toplam[kat] = kategori_toplam.get(kat, 0) + h.get('tutar', 0)
        for r, (kat, toplam_k) in enumerate(sorted(kategori_toplam.items(), key=lambda x: x[0]), 2):
            ws_kat.cell(row=r, column=1, value=kat)
            ws_kat.cell(row=r, column=2, value=toplam_k)
            ws_kat.cell(row=r, column=1).border = border
            ws_kat.cell(row=r, column=2).border = border
        ws_kat.column_dimensions['A'].width = 20
        ws_kat.column_dimensions['B'].width = 15

        # ===== İSTATİSTİKLER SAYFASI =====
        ws_istat = wb.create_sheet("İstatistikler")
        ws_istat.cell(row=1, column=1, value="Kategori").font = header_font
        ws_istat.cell(row=1, column=2, value="Toplam").font = header_font
        ws_istat.cell(row=1, column=3, value="Ortalama").font = header_font
        ws_istat.cell(row=1, column=4, value="En Küçük").font = header_font
        ws_istat.cell(row=1, column=5, value="En Büyük").font = header_font
        from collections import defaultdict
        cat_vals = defaultdict(list)
        for h in veri_export:
            cat_vals[h.get('kategori', 'Diğer')].append(h.get('tutar', 0))
        r = 2
        for kat in sorted(cat_vals.keys()):
            vals = cat_vals[kat]
            toplam_k = sum(vals)
            ort = toplam_k / len(vals) if vals else 0
            mn = min(vals) if vals else 0
            mx = max(vals) if vals else 0
            ws_istat.cell(row=r, column=1, value=kat)
            ws_istat.cell(row=r, column=2, value=toplam_k)
            ws_istat.cell(row=r, column=3, value=ort)
            ws_istat.cell(row=r, column=4, value=mn)
            ws_istat.cell(row=r, column=5, value=mx)
            for col in range(1,6):
                ws_istat.cell(row=r, column=col).border = border
            r += 1
        ws_istat.column_dimensions['A'].width = 20
        ws_istat.column_dimensions['B'].width = 15
        ws_istat.column_dimensions['C'].width = 12
        ws_istat.column_dimensions['D'].width = 12
        ws_istat.column_dimensions['E'].width = 12

        # ===== ÖZETİ SAYFASI =====
        ws_ozet = wb.create_sheet("Özet", 0)
        
        ws_ozet.cell(row=1, column=1, value="BÜTÇE TRAKİ ÖZETİ").font = Font(bold=True, size=14)
        
        row = 3
        ws_ozet.cell(row=row, column=1, value="Toplam Harcama").font = Font(bold=True)
        ws_ozet.cell(row=row, column=2, value=sum(h['tutar'] for h in veri_export))
        
        row += 1
        ws_ozet.cell(row=row, column=1, value="Toplam Gelir").font = Font(bold=True)
        ws_ozet.cell(row=row, column=2, value=sum(g['tutar'] for g in gelirler))
        
        row += 1
        toplam_gelir = sum(g['tutar'] for g in gelirler)
        toplam_gider = sum(h['tutar'] for h in veri_export)
        fark = toplam_gelir - toplam_gider
        ws_ozet.cell(row=row, column=1, value="Net Fark").font = Font(bold=True)
        ws_ozet.cell(row=row, column=2, value=fark).font = Font(bold=True, size=12)
        
        row += 1
        ws_ozet.cell(row=row, column=1, value="Kumbara Bakiyesi").font = Font(bold=True)
        ws_ozet.cell(row=row, column=2, value=kumbara_bakiye).font = Font(bold=True)
        
        row += 1
        ws_ozet.cell(row=row, column=1, value="Belirlenen Bütçe").font = Font(bold=True)
        ws_ozet.cell(row=row, column=2, value=butce)
        
        ws_ozet.column_dimensions['A'].width = 20
        ws_ozet.column_dimensions['B'].width = 15
        
        # Excel dosyasını kaydet
        dosya_adi = f"butce_takipcisi_{datetime.now().strftime('%d_%m_%Y_%H_%M')}.xlsx"
        wb.save(dosya_adi)
        
        print(f"✓ Veriler '{dosya_adi}' dosyasına başarıyla aktarıldı!")
        print(f"  📊 5 sayfa oluşturuldu: Özet, Harcamalar, Gelirler, Kumbara")
    except Exception as e:
        print(f"❌ Excel export sırasında hata oluştu: {e}")

def grafik_goster():
    """Grafik ve chart gösterir"""
    if not harcamalar:
        print("\n❌ Grafik oluşturmak için harcama kaydı gereklidir.")
        return
    
    print("\n--- GRAFIK SEÇENEKLERİ ---")
    print("1. Kategori Dağılımı (Pie Chart)")
    print("2. Kategori Harcamaları (Bar Chart)")
    print("3. Zaman İçinde Harcama Trendi (Line Chart)")
    secim = input("Seçiminiz (1-3): ")
    
    if secim == "1":
        # Pie Chart - Kategori dağılımı
        kategori_toplamı = {}
        for harcama in harcamalar:
            kat = harcama['kategori']
            if kat not in kategori_toplamı:
                kategori_toplamı[kat] = 0
            kategori_toplamı[kat] += harcama['tutar']
        
        plt.figure(figsize=(10, 6))
        plt.pie(kategori_toplamı.values(), labels=kategori_toplamı.keys(), autopct='%1.1f%%', startangle=90)
        plt.title('Harcama Kategorileri Dağılımı')
        plt.tight_layout()
        plt.show()
        
    elif secim == "2":
        # Bar Chart - Kategori harcamaları
        kategori_toplamı = {}
        for harcama in harcamalar:
            kat = harcama['kategori']
            if kat not in kategori_toplamı:
                kategori_toplamı[kat] = 0
            kategori_toplamı[kat] += harcama['tutar']
        
        kategoriler = sorted(kategori_toplamı.keys())
        tutarlar = [kategori_toplamı[k] for k in kategoriler]
        
        plt.figure(figsize=(10, 6))
        plt.bar(kategoriler, tutarlar, color='skyblue', edgecolor='navy', alpha=0.7)
        plt.xlabel('Kategoriler', fontsize=12)
        plt.ylabel('Harcama (TL)', fontsize=12)
        plt.title('Kategorilere Göre Toplam Harcamalar', fontsize=14)
        plt.xticks(rotation=45, ha='right')
        plt.grid(axis='y', alpha=0.3)
        
        # Değerleri çubuğun üstüne yaz
        for i, v in enumerate(tutarlar):
            plt.text(i, v + 5, str(v), ha='center', va='bottom', fontweight='bold')
        
        plt.tight_layout()
        plt.show()
        
    elif secim == "3":
        # Line Chart - Tarih bazında harcama trendi
        harcamalar_siralı = sorted(harcamalar, key=lambda x: datetime.strptime(x['tarih'], "%d.%m.%Y"))
        
        tarihler = []
        kumulatif = 0
        kumulatif_tutarlar = []
        
        for harcama in harcamalar_siralı:
            tarihler.append(harcama['tarih'])
            kumulatif += harcama['tutar']
            kumulatif_tutarlar.append(kumulatif)
        
        plt.figure(figsize=(12, 6))
        plt.plot(range(len(tarihler)), kumulatif_tutarlar, marker='o', linestyle='-', linewidth=2, markersize=4, color='green')
        plt.xlabel('Tarih Sırası', fontsize=12)
        plt.ylabel('Birikmiş Harcama (TL)', fontsize=12)
        plt.title('Harcama Trendi (Zaman İçinde Birikmiş)', fontsize=14)
        plt.grid(True, alpha=0.3)
        
        # Her 5. tarihi göster
        tick_positions = list(range(0, len(tarihler), max(1, len(tarihler)//10)))
        tick_labels = [tarihler[i] if i < len(tarihler) else '' for i in tick_positions]
        plt.xticks(tick_positions, tick_labels, rotation=45, ha='right')
        
        plt.tight_layout()
        plt.show()
    else:
        print("❌ Geçersiz seçim!")

def karsilastir():
    """İki dönem arasında harcamaları karşılaştırır"""
    if len(harcamalar) < 2:
        print("\n❌ Karşılaştırma için yeterli harcama kaydı yok.")
        return
    
    print("\n--- DÖNEM KARŞILAŞTIRMASI ---")
    
    # İlk dönem
    print("\n📅 BİRİNCİ DÖNEMİ SEÇİN:")
    while True:
        basla1_str = input("Başlangıç (GG.AA.YYYY): ").strip()
        try:
            basla1 = datetime.strptime(basla1_str, "%d.%m.%Y")
            break
        except ValueError:
            print("❌ Geçersiz tarih!")
    
    while True:
        bitis1_str = input("Bitiş (GG.AA.YYYY): ").strip()
        try:
            bitis1 = datetime.strptime(bitis1_str, "%d.%m.%Y")
            break
        except ValueError:
            print("❌ Geçersiz tarih!")
    
    # İkinci dönem
    print("\n📅 İKİNCİ DÖNEMİ SEÇİN:")
    while True:
        basla2_str = input("Başlangıç (GG.AA.YYYY): ").strip()
        try:
            basla2 = datetime.strptime(basla2_str, "%d.%m.%Y")
            break
        except ValueError:
            print("❌ Geçersiz tarih!")
    
    while True:
        bitis2_str = input("Bitiş (GG.AA.YYYY): ").strip()
        try:
            bitis2 = datetime.strptime(bitis2_str, "%d.%m.%Y")
            break
        except ValueError:
            print("❌ Geçersiz tarih!")
    
    # Dönemlere göre harcamaları filtrele
    donem1 = []
    donem2 = []
    
    for harcama in harcamalar:
        h_tarih = datetime.strptime(harcama['tarih'], "%d.%m.%Y")
        if basla1 <= h_tarih <= bitis1:
            donem1.append(harcama)
        if basla2 <= h_tarih <= bitis2:
            donem2.append(harcama)
    
    if not donem1 or not donem2:
        print("\n❌ Seçilen dönemlerde harcama bulunamadı.")
        return
    
    # Kategori bazlı hesaplama
    kat1 = {}
    kat2 = {}
    
    for h in donem1:
        k = h['kategori']
        if k not in kat1:
            kat1[k] = 0
        kat1[k] += h['tutar']
    
    for h in donem2:
        k = h['kategori']
        if k not in kat2:
            kat2[k] = 0
        kat2[k] += h['tutar']
    
    # Raporlama
    print(f"\n===== KARŞILAŞTIRMA ({basla1_str} - {bitis1_str}) vs ({basla2_str} - {bitis2_str}) =====")
    
    toplam1 = sum(kat1.values())
    toplam2 = sum(kat2.values())
    
    print(f"\n💰 Dönem 1 Toplam: {toplam1} TL ({len(donem1)} işlem)")
    print(f"💰 Dönem 2 Toplam: {toplam2} TL ({len(donem2)} işlem)")
    
    fark = toplam2 - toplam1
    yuzde_degisim = (fark / toplam1 * 100) if toplam1 > 0 else 0
    
    if fark > 0:
        print(f"\n📈 Artış: +{fark} TL (+%{yuzde_degisim:.1f})")
    elif fark < 0:
        print(f"\n📉 Azalış: {fark} TL ({yuzde_degisim:.1f}%)")
    else:
        print(f"\n➡️  Değişim yok")
    
    # Kategori bazlı karşılaştırma
    print("\n📂 KATEGORİ BAZLI KARŞILAŞTIRMA:")
    print(f"{'Kategori':<15} {'Dönem 1':<12} {'Dönem 2':<12} {'Fark':<12} {'Değişim':<10}")
    print("-" * 65)
    
    tum_kategoriler = set(list(kat1.keys()) + list(kat2.keys()))
    
    for kat in sorted(tum_kategoriler):
        d1 = kat1.get(kat, 0)
        d2 = kat2.get(kat, 0)
        fark_kat = d2 - d1
        
        if d1 > 0:
            yuzde_kat = (fark_kat / d1 * 100)
            yuzde_str = f"{yuzde_kat:+.1f}%"
        elif d2 > 0:
            yuzde_str = "+∞"
        else:
            yuzde_str = "—"
        
        fark_str = f"{fark_kat:+.0f}"
        print(f"{kat:<15} {d1:<12.0f} {d2:<12.0f} {fark_str:<12} {yuzde_str:<10}")

def butce_tahmini():
    """Ay sonu bütçe tahmini yapar"""
    if not harcamalar:
        print("❌ Henüz harcama kaydı yok.")
        return
    
    # Bugünün tarihi
    bugün = datetime.now()
    ay_adi = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
              "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
    
    print(f"\n📊 {ay_adi[bugün.month-1]} {bugün.year} BÜTÇE TAHMİNİ")
    print("=" * 50)
    
    # Ay başı ve ay sonu tarihlerini belirle
    ay_baslangici = datetime(bugün.year, bugün.month, 1)
    if bugün.month == 12:
        ay_sonu = datetime(bugün.year + 1, 1, 1)
    else:
        ay_sonu = datetime(bugün.year, bugün.month + 1, 1)
    
    # Ay sonu tarihi (31, 30, 29, 28 vs)
    import calendar
    ayın_gün_sayısı = calendar.monthrange(bugün.year, bugün.month)[1]
    
    # Bu ayın harcamalarını filtrele
    bu_ay_harcamalar = []
    for h in harcamalar:
        tarih = datetime.strptime(h['tarih'], "%d.%m.%Y")
        if tarih.year == bugün.year and tarih.month == bugün.month:
            bu_ay_harcamalar.append(h)
    
    # Hesaplamalar
    toplam_harcama = sum(h['tutar'] for h in bu_ay_harcamalar)
    geçen_gün = (bugün - ay_baslangici).days + 1  # Bugün dahil
    kalan_gün = ayın_gün_sayısı - geçen_gün
    
    if geçen_gün > 0:
        günlük_ortalama = toplam_harcama / geçen_gün
    else:
        günlük_ortalama = 0
    
    # Tahmin
    tahmini_kalan_harcama = günlük_ortalama * kalan_gün
    tahmini_ay_sonu = toplam_harcama + tahmini_kalan_harcama
    
    # Bütçe kontrolü
    eğer_bütçe_var = butce > 0
    
    print(f"\n📅 Tarih: {bugün.strftime('%d.%m.%Y')}")
    print(f"📈 Ayın Günü: {geçen_gün}/{ayın_gün_sayısı}")
    print(f"💸 Bu Aya Kadar Harcanan: {toplam_harcama:.2f} TL")
    print(f"📊 Günlük Ortalama: {günlük_ortalama:.2f} TL/gün")
    print(f"⏳ Kalan Gün: {kalan_gün}")
    print(f"🔮 Tahmini Kalan Harcama: {tahmini_kalan_harcama:.2f} TL")
    print(f"📌 TAHMINI AY SONU TOPLAMI: {tahmini_ay_sonu:.2f} TL")
    
    if eğer_bütçe_var:
        print(f"\n💰 Belirlenen Aylık Bütçe: {butce:.2f} TL")
        fark = butce - tahmini_ay_sonu
        yuzde = (fark / butce) * 100
        
        if fark >= 0:
            print(f"✅ Bütçeye Kalan: {fark:.2f} TL ({yuzde:.1f}%)")
        else:
            print(f"⚠️ BÜTÇE AŞIMININIZ: {abs(fark):.2f} TL ({abs(yuzde):.1f}%)")
    
    print("\nℹ️ Not: Bu tahmin, mevcut harcama hızınızın ay sonuna kadar")
    print("devam edeceği varsayımına dayanır.")

def butce_menu():
    """Bütçe yönetimi menüsü"""
    while True:
        print("\n" + "=" * 40)
        print("1. Bütçe Belirle")
        print("2. Bütçe Durumunu Görüntüle")
        print("3. Geri Dön")
        print("=" * 40)
        secim = input("Seçiminiz: ")
        
        if secim == "1":
            butce_belirle()
        elif secim == "2":
            butce_goster()
        elif secim == "3":
            break
        else:
            print("❌ Geçersiz seçim!")

def menu_goster():
    """Ana menüyü gösterir"""
    print("\n" + "=" * 40)
    print("1. Harcama Ekle")
    print("2. Tüm Harcamaları Görüntüle")
    print("3. Kategori Bazlı Harcamaları Gör")
    print("4. İstatistikleri Gör")
    print("5. Tarih Aralığına Göre Filtrele")
    print("6. Bütçe Yönetimi")
    print("7. Tekrarlayan Harcamalar")
    print("8. Dönem Karşılaştırması")
    print("9. Bütçe Tahmini")
    print("10. Gelir Ekle")
    print("11. Tüm Gelirleri Görüntüle")
    print("12. Gider vs Gelir Özeti")
    print("13. Kumbara Yönetimi")
    print("14. Grafik Gör")
    print("15. CSV'ye Aktar")
    print("16. Excel'e Aktar")
    print("17. Harcama Sil")
    print("18. Harcama Düzenle")
    print("19. Çıkış")
    print("=" * 40)

# Ana program döngüsü
veri_yukle()

while True:
    menu_goster()
    
    # Kumbara otomatik güncelleme
    if kumbara_ayari["mod"] is not None:
        kumbara_ekle()
    
    secim = input("\nSeçiminizi yapınız (1-19): ")
    
    if secim == "1":
        harcama_ekle()
    elif secim == "2":
        harcamaları_goster()
    elif secim == "3":
        kategori_bazli_goster()
    elif secim == "4":
        istatistikler_goster()
    elif secim == "5":
        tarih_araligi_goster()
    elif secim == "6":
        butce_menu()
    elif secim == "7":
        tekrarlayan_menu()
    elif secim == "8":
        karsilastir()
    elif secim == "9":
        butce_tahmini()
    elif secim == "10":
        gelir_ekle()
    elif secim == "11":
        gelir_goster()
    elif secim == "12":
        gider_gelir_ozeti()
    elif secim == "13":
        kumbara_menu()
    elif secim == "14":
        grafik_goster()
    elif secim == "15":
        csv_export()
    elif secim == "16":
        excel_export()
    elif secim == "17":
        harcama_sil()
    elif secim == "18":
        harcama_duzenle()
    elif secim == "19":
        print("\nÇıkılıyor... Hoşça kalın!")
        veri_kaydet()
        gelir_kaydet()
        kumbara_kaydet()
        butce_kaydet()
        tekrarlayan_kaydet()
        break
    else:
        print("❌ Geçersiz seçim! Lütfen 1-19 arasında seçim yapınız.")
