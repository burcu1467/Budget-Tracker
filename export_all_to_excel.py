from datetime import datetime
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Dosya adları (aynı olduğu varsayılıyor)
DOSYA_ADI = "harcamalar.json"
GELIR_DOSYASI = "gelirler.json"
TEKRARLAYAN_DOSYASI = "tekrarlayan.json"
BUTCE_DOSYASI = "butce.json"
KUMBARA_DOSYASI = "kumbara.json"

# Yükle yardımcı
def load_json(path):
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"⚠️ {path} okunamadı: {e}")
    return []

harcamalar = load_json(DOSYA_ADI)
gelirler = load_json(GELIR_DOSYASI)
tekrarlayan_harcamalar = load_json(TEKRARLAYAN_DOSYASI)
butce = 0
if os.path.exists(BUTCE_DOSYASI):
    try:
        with open(BUTCE_DOSYASI, "r", encoding="utf-8") as f:
            butce = json.load(f)
    except:
        butce = 0
kumbara_bakiye = 0
kumbara_islemleri = []
kumbara_ayari = {"gunluk_tutar":0, "haftalik_tutar":0, "mod": None, 'son_tarih': None}
if os.path.exists(KUMBARA_DOSYASI):
    try:
        with open(KUMBARA_DOSYASI, "r", encoding="utf-8") as f:
            kd = json.load(f)
            kumbara_bakiye = kd.get('bakiye', 0)
            kumbara_islemleri = kd.get('islemler', [])
            kumbara_ayari = kd.get('ayar', kumbara_ayari)
    except Exception as e:
        print(f"⚠️ {KUMBARA_DOSYASI} okunamadı: {e}")

# Oluştur
wb = Workbook()
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal="center", vertical="center")

# Harcamalar
ws_harcama = wb.active
ws_harcama.title = "Harcamalar"
headers = ["Tarih", "Kategori", "Açıklama", "Tutar (TL)"]
for col, header in enumerate(headers,1):
    cell = ws_harcama.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = center_align
for row, h in enumerate(harcamalar,2):
    ws_harcama.cell(row=row, column=1, value=h.get('tarih'))
    ws_harcama.cell(row=row, column=2, value=h.get('kategori'))
    ws_harcama.cell(row=row, column=3, value=h.get('aciklama'))
    ws_harcama.cell(row=row, column=4, value=h.get('tutar'))
    for col in range(1,5):
        ws_harcama.cell(row=row, column=col).border = border
toplam_row = len(harcamalar) + 2
ws_harcama.cell(row=toplam_row, column=1, value="TOPLAM")
ws_harcama.cell(row=toplam_row, column=4, value=sum(h.get('tutar',0) for h in harcamalar))
for col in range(1,5):
    cell = ws_harcama.cell(row=toplam_row, column=col)
    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.border = border
ws_harcama.column_dimensions['A'].width = 12
ws_harcama.column_dimensions['B'].width = 15
ws_harcama.column_dimensions['C'].width = 25
ws_harcama.column_dimensions['D'].width = 12

# Gelirler
ws_gelir = wb.create_sheet("Gelirler")
headers_gelir = ["Tarih", "Kaynağı", "Tutar (TL)"]
for col, header in enumerate(headers_gelir,1):
    cell = ws_gelir.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = center_align
for row, g in enumerate(gelirler,2):
    ws_gelir.cell(row=row, column=1, value=g.get('tarih'))
    ws_gelir.cell(row=row, column=2, value=g.get('aciklama'))
    ws_gelir.cell(row=row, column=3, value=g.get('tutar'))
    for col in range(1,4):
        ws_gelir.cell(row=row, column=col).border = border
if gelirler:
    gelir_toplam_row = len(gelirler) + 2
    ws_gelir.cell(row=gelir_toplam_row, column=1, value="TOPLAM")
    ws_gelir.cell(row=gelir_toplam_row, column=3, value=sum(g.get('tutar',0) for g in gelirler))
    for col in range(1,4):
        cell = ws_gelir.cell(row=gelir_toplam_row, column=col)
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = border
ws_gelir.column_dimensions['A'].width = 12
ws_gelir.column_dimensions['B'].width = 20
ws_gelir.column_dimensions['C'].width = 12

# Kumbara
ws_kumbara = wb.create_sheet("Kumbara")
ws_kumbara.cell(row=1, column=1, value="Kumbara Bakiyesi").font = Font(bold=True, size=12)
ws_kumbara.cell(row=1, column=2, value=kumbara_bakiye).font = Font(size=12)
ws_kumbara.cell(row=3, column=1, value="Tasarruf Modu").font = Font(bold=True)
mod_text = (f"Günlük: {kumbara_ayari.get('gunluk_tutar')} TL" if kumbara_ayari.get('mod') == 'gunluk' else f"Haftalık: {kumbara_ayari.get('haftalik_tutar')} TL")
ws_kumbara.cell(row=3, column=2, value=mod_text)
if kumbara_islemleri:
    ws_kumbara.cell(row=5, column=1, value="İşlem Tarihi").font = Font(bold=True)
    ws_kumbara.cell(row=5, column=2, value="Tür").font = Font(bold=True)
    ws_kumbara.cell(row=5, column=3, value="Tutar").font = Font(bold=True)
    for row, islem in enumerate(kumbara_islemleri,6):
        ws_kumbara.cell(row=row, column=1, value=islem.get('tarih'))
        ws_kumbara.cell(row=row, column=2, value=islem.get('tür'))
        ws_kumbara.cell(row=row, column=3, value=islem.get('tutar'))
ws_kumbara.column_dimensions['A'].width = 15
ws_kumbara.column_dimensions['B'].width = 20
ws_kumbara.column_dimensions['C'].width = 12

# Tekrarlayan
ws_tekr = wb.create_sheet("Tekrarlayan")
headers_tekr = ["Gün", "Kategori", "Açıklama", "Tutar", "Aktif"]
for col, header in enumerate(headers_tekr,1):
    cell = ws_tekr.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = center_align
for row, t in enumerate(tekrarlayan_harcamalar,2):
    ws_tekr.cell(row=row, column=1, value=t.get('gun'))
    ws_tekr.cell(row=row, column=2, value=t.get('kategori'))
    ws_tekr.cell(row=row, column=3, value=t.get('aciklama'))
    ws_tekr.cell(row=row, column=4, value=t.get('tutar'))
    ws_tekr.cell(row=row, column=5, value=('Evet' if t.get('aktif') else 'Hayır'))
    for col in range(1,6):
        ws_tekr.cell(row=row, column=col).border = border
ws_tekr.column_dimensions['A'].width = 8
ws_tekr.column_dimensions['B'].width = 15
ws_tekr.column_dimensions['C'].width = 30
ws_tekr.column_dimensions['D'].width = 12
ws_tekr.column_dimensions['E'].width = 8

# Kategori özeti
ws_kat = wb.create_sheet("Kategori Özeti")
ws_kat.cell(row=1, column=1, value="Kategori").font = header_font
ws_kat.cell(row=1, column=2, value="Toplam Harcama").font = header_font
kategori_toplam = {}
for h in harcamalar:
    kat = h.get('kategori', 'Diğer')
    kategori_toplam[kat] = kategori_toplam.get(kat,0) + h.get('tutar',0)
for r, (kat, toplam_k) in enumerate(sorted(kategori_toplam.items(), key=lambda x: x[0]),2):
    ws_kat.cell(row=r, column=1, value=kat)
    ws_kat.cell(row=r, column=2, value=toplam_k)
    ws_kat.cell(row=r, column=1).border = border
    ws_kat.cell(row=r, column=2).border = border
ws_kat.column_dimensions['A'].width = 20
ws_kat.column_dimensions['B'].width = 15

# İstatistikler
from collections import defaultdict
ws_istat = wb.create_sheet("İstatistikler")
ws_istat.cell(row=1, column=1, value="Kategori").font = header_font
ws_istat.cell(row=1, column=2, value="Toplam").font = header_font
ws_istat.cell(row=1, column=3, value="Ortalama").font = header_font
ws_istat.cell(row=1, column=4, value="En Küçük").font = header_font
ws_istat.cell(row=1, column=5, value="En Büyük").font = header_font
cat_vals = defaultdict(list)
for h in harcamalar:
    cat_vals[h.get('kategori','Diğer')].append(h.get('tutar',0))
r = 2
for kat in sorted(cat_vals.keys()):
    vals = cat_vals[kat]
    toplam_k = sum(vals)
    ort = toplam_k / len(vals) if vals else 0
    mn = min(vals) if vals else 0
    mx = max(vals) if vals else 0
    ws_istat.cell(row=r, column=1, value=kat)
    ws_istat.cell(row=r, column=2, value=toplam_k)
    ws_istat.cell(row=r, column=3, value=ort)
    ws_istat.cell(row=r, column=4, value=mn)
    ws_istat.cell(row=r, column=5, value=mx)
    for col in range(1,6):
        ws_istat.cell(row=r, column=col).border = border
    r += 1
ws_istat.column_dimensions['A'].width = 20
ws_istat.column_dimensions['B'].width = 15
ws_istat.column_dimensions['C'].width = 12
ws_istat.column_dimensions['D'].width = 12
ws_istat.column_dimensions['E'].width = 12

# Özet
ws_ozet = wb.create_sheet("Özet", 0)
ws_ozet.cell(row=1, column=1, value="BÜTÇE TAKİPÇİSİ ÖZETİ").font = Font(bold=True, size=14)
row = 3
ws_ozet.cell(row=row, column=1, value="Toplam Harcama").font = Font(bold=True)
ws_ozet.cell(row=row, column=2, value=sum(h.get('tutar',0) for h in harcamalar))
row += 1
ws_ozet.cell(row=row, column=1, value="Toplam Gelir").font = Font(bold=True)
ws_ozet.cell(row=row, column=2, value=sum(g.get('tutar',0) for g in gelirler))
row += 1
toplam_gelir = sum(g.get('tutar',0) for g in gelirler)
toplam_gider = sum(h.get('tutar',0) for h in harcamalar)
fark = toplam_gelir - toplam_gider
ws_ozet.cell(row=row, column=1, value="Net Fark").font = Font(bold=True)
ws_ozet.cell(row=row, column=2, value=fark).font = Font(bold=True, size=12)
row += 1
ws_ozet.cell(row=row, column=1, value="Kumbara Bakiyesi").font = Font(bold=True)
ws_ozet.cell(row=row, column=2, value=kumbara_bakiye).font = Font(bold=True)
row += 1
ws_ozet.cell(row=row, column=1, value="Belirlenen Bütçe").font = Font(bold=True)
ws_ozet.cell(row=row, column=2, value=butce)
ws_ozet.column_dimensions['A'].width = 20
ws_ozet.column_dimensions['B'].width = 15

dosya_adi = f"butce_takipcisi_{datetime.now().strftime('%d_%m_%Y_%H_%M_%S')}.xlsx"
wb.save(dosya_adi)
print(f"✓ Excel oluşturuldu: {dosya_adi}")
